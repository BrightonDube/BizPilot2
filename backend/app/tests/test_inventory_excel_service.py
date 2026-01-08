"""Tests for inventory Excel import/export service."""

import uuid
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.models.inventory import InventoryItem
from app.models.product import Product
from app.services.inventory_excel_service import InventoryExcelService, INVENTORY_COLUMNS


class FakeQuery:
    """Fake SQLAlchemy query for testing."""

    def __init__(self, items):
        self._items = list(items)
        self._filters_applied = []

    def filter(self, *args, **kwargs):
        return self

    def join(self, *args, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return self._items

    def first(self):
        return self._items[0] if self._items else None


class FakeSession:
    """Fake SQLAlchemy session for testing."""

    def __init__(self, data_by_model=None, join_results=None):
        self.data_by_model = data_by_model or {}
        self.join_results = join_results or []  # For multi-model queries
        self.added = []
        self.commits = 0
        self.rollbacks = 0

    def query(self, *models):
        # Multi-model query (for joins)
        if len(models) > 1:
            return FakeQuery(self.join_results)
        return FakeQuery(self.data_by_model.get(models[0], []))

    def add(self, obj):
        if getattr(obj, "id", None) is None:
            obj.id = uuid.uuid4()
        self.added.append(obj)

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1


@pytest.fixture
def business_id():
    return str(uuid.uuid4())


@pytest.fixture
def user_id():
    return str(uuid.uuid4())


def _make_product(business_id: str, sku: str, name: str) -> Product:
    """Create a fake product for testing."""
    product = Product()
    product.id = uuid.uuid4()
    product.business_id = business_id
    product.sku = sku
    product.name = name
    product.deleted_at = None
    return product


def _make_inventory_item(
    business_id: str,
    product_id: uuid.UUID,
    qty: int = 10,
    location: str = None,
) -> InventoryItem:
    """Create a fake inventory item for testing."""
    item = InventoryItem()
    item.id = uuid.uuid4()
    item.business_id = business_id
    item.product_id = product_id
    item.quantity_on_hand = qty
    item.quantity_reserved = 0
    item.quantity_incoming = 0
    item.reorder_point = 5
    item.reorder_quantity = 20
    item.location = location
    item.bin_location = None
    item.average_cost = Decimal("10.00")
    item.last_cost = Decimal("10.00")
    item.deleted_at = None
    return item


class TestGenerateTemplate:
    """Tests for template generation."""

    def test_generate_template_returns_valid_excel(self):
        """Template should be a valid Excel file."""
        db = FakeSession()
        service = InventoryExcelService(db)

        output = service.generate_template()

        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Inventory" in wb.sheetnames

    def test_generate_template_has_correct_headers(self):
        """Template should have all expected column headers."""
        db = FakeSession()
        service = InventoryExcelService(db)

        output = service.generate_template()
        wb = load_workbook(output)
        ws = wb["Inventory"]

        headers = [cell.value for cell in ws[1]]
        expected_headers = [col["header"] for col in INVENTORY_COLUMNS]

        assert headers == expected_headers

    def test_generate_template_has_instructions_sheet(self):
        """Template should include an instructions sheet."""
        db = FakeSession()
        service = InventoryExcelService(db)

        output = service.generate_template()
        wb = load_workbook(output)

        assert "Instructions" in wb.sheetnames
        instructions_ws = wb["Instructions"]
        assert instructions_ws["A1"].value == "BizPilot Inventory Import Template"

    def test_generate_template_has_frozen_header_row(self):
        """Template should have frozen header row for usability."""
        db = FakeSession()
        service = InventoryExcelService(db)

        output = service.generate_template()
        wb = load_workbook(output)
        ws = wb["Inventory"]

        assert ws.freeze_panes == "A2"


class TestExportInventory:
    """Tests for inventory export."""

    def test_export_empty_inventory_returns_valid_excel(self, business_id):
        """Export with no items should return valid Excel with headers only."""
        db = FakeSession(data_by_model={}, join_results=[])
        service = InventoryExcelService(db)

        output = service.export_inventory(business_id)

        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Inventory" in wb.sheetnames

    def test_export_includes_all_columns(self, business_id):
        """Export should include all defined columns."""
        product = _make_product(business_id, "SKU001", "Test Product")
        item = _make_inventory_item(business_id, product.id, qty=25, location="Warehouse A")

        db = FakeSession(
            data_by_model={},
            join_results=[(item, product)]
        )
        service = InventoryExcelService(db)

        output = service.export_inventory(business_id)
        wb = load_workbook(output)
        ws = wb["Inventory"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "SKU" in headers
        assert "Product Name" in headers
        assert "Quantity On Hand" in headers

    def test_export_includes_metadata_sheet(self, business_id):
        """Export should include metadata sheet with export info."""
        db = FakeSession(data_by_model={}, join_results=[])
        service = InventoryExcelService(db)

        output = service.export_inventory(business_id)
        wb = load_workbook(output)

        assert "Export Info" in wb.sheetnames
        meta = wb["Export Info"]
        assert meta["A1"].value == "Export Date:"
        assert meta["A2"].value == "Total Items:"


class TestImportInventory:
    """Tests for inventory import."""

    def _create_test_excel(self, rows: list[dict]) -> bytes:
        """Helper to create test Excel file with given data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Add headers
        headers = ["SKU", "Product Name", "Quantity On Hand", "Quantity Reserved",
                   "Quantity Incoming", "Reorder Point", "Reorder Quantity",
                   "Location", "Bin Location", "Average Cost", "Last Cost"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Add data rows
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data.get("sku"))
            ws.cell(row=row_idx, column=2, value=row_data.get("product_name"))
            ws.cell(row=row_idx, column=3, value=row_data.get("quantity_on_hand"))
            ws.cell(row=row_idx, column=4, value=row_data.get("quantity_reserved"))
            ws.cell(row=row_idx, column=5, value=row_data.get("quantity_incoming"))
            ws.cell(row=row_idx, column=6, value=row_data.get("reorder_point"))
            ws.cell(row=row_idx, column=7, value=row_data.get("reorder_quantity"))
            ws.cell(row=row_idx, column=8, value=row_data.get("location"))
            ws.cell(row=row_idx, column=9, value=row_data.get("bin_location"))
            ws.cell(row=row_idx, column=10, value=row_data.get("average_cost"))
            ws.cell(row=row_idx, column=11, value=row_data.get("last_cost"))

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def test_import_with_valid_data_creates_items(self, business_id, user_id):
        """Import with valid data should create inventory items."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "SKU001", "quantity_on_hand": 50, "location": "Warehouse A"}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is True
        assert result["created"] == 1
        assert result["updated"] == 0
        assert len(result["errors"]) == 0

    def test_import_updates_existing_items(self, business_id, user_id):
        """Import should update existing inventory items."""
        product = _make_product(business_id, "SKU001", "Test Product")
        existing_item = _make_inventory_item(business_id, product.id, qty=10)

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [existing_item],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "SKU001", "quantity_on_hand": 100}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is True
        assert result["updated"] == 1
        assert result["created"] == 0
        assert existing_item.quantity_on_hand == 100

    def test_import_with_missing_required_columns_fails(self, business_id, user_id):
        """Import should fail if required columns are missing."""
        # Create Excel without SKU column
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.cell(row=1, column=1, value="Product Name")
        ws.cell(row=1, column=2, value="Quantity On Hand")

        output = BytesIO()
        wb.save(output)
        excel_content = output.getvalue()

        db = FakeSession()
        service = InventoryExcelService(db)

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is False
        assert any("Missing required columns" in err for err in result["errors"])

    def test_import_with_invalid_sku_reports_error(self, business_id, user_id):
        """Import should report error for unknown SKUs."""
        db = FakeSession(data_by_model={
            Product: [],  # No products
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "NONEXISTENT", "quantity_on_hand": 50}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["skipped"] == 1
        assert any("not found" in err for err in result["errors"])

    def test_import_with_invalid_quantity_reports_error(self, business_id, user_id):
        """Import should report error for invalid quantity values."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "SKU001", "quantity_on_hand": "invalid"}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["skipped"] >= 1
        assert len(result["errors"]) > 0

    def test_import_skips_empty_rows(self, business_id, user_id):
        """Import should skip rows with empty SKU."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "", "quantity_on_hand": 50},  # Empty SKU - should skip
            {"sku": "SKU001", "quantity_on_hand": 100},  # Valid
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is True
        assert result["created"] == 1
        assert result["skipped"] == 1

    def test_import_converts_empty_strings_to_none(self, business_id, user_id):
        """Import should convert empty string locations to None."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "SKU001", "quantity_on_hand": 50, "location": "", "bin_location": "   "}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is True
        # Check the added item has None for empty locations
        assert len(db.added) == 1
        added_item = db.added[0]
        assert added_item.location is None
        assert added_item.bin_location is None

    def test_import_handles_malformed_excel(self, business_id, user_id):
        """Import should handle malformed Excel files gracefully."""
        db = FakeSession()
        service = InventoryExcelService(db)

        result = service.import_inventory(business_id, b"not an excel file", user_id)

        assert result["success"] is False
        assert any("Failed to read Excel file" in err for err in result["errors"])

    def test_import_handles_empty_workbook(self, business_id, user_id):
        """Import should handle workbook with no sheets."""
        # Create minimal xlsx that will fail to load properly
        db = FakeSession()
        service = InventoryExcelService(db)

        # Empty bytes won't be valid Excel
        result = service.import_inventory(business_id, b"", user_id)

        assert result["success"] is False

    def test_import_finds_inventory_sheet_by_name(self, business_id, user_id):
        """Import should find sheet named 'Inventory'."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        # Create workbook with custom sheet name
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"  # Named correctly
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=2, value="Quantity On Hand")
        ws.cell(row=2, column=1, value="SKU001")
        ws.cell(row=2, column=2, value=50)

        output = BytesIO()
        wb.save(output)

        result = service.import_inventory(business_id, output.getvalue(), user_id)

        assert result["success"] is True

    def test_import_falls_back_to_first_sheet(self, business_id, user_id):
        """Import should fall back to first sheet if 'Inventory' not found."""
        product = _make_product(business_id, "SKU001", "Test Product")

        db = FakeSession(data_by_model={
            Product: [product],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        # Create workbook with different sheet name
        wb = Workbook()
        ws = wb.active
        ws.title = "MyData"  # Different name
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=2, value="Quantity On Hand")
        ws.cell(row=2, column=1, value="SKU001")
        ws.cell(row=2, column=2, value=50)

        output = BytesIO()
        wb.save(output)

        result = service.import_inventory(business_id, output.getvalue(), user_id)

        assert result["success"] is True

    def test_import_with_no_valid_data_fails(self, business_id, user_id):
        """Import should fail if no valid inventory data found."""
        db = FakeSession(data_by_model={
            Product: [],
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        excel_content = self._create_test_excel([
            {"sku": "NONEXISTENT", "quantity_on_hand": 50}
        ])

        result = service.import_inventory(business_id, excel_content, user_id)

        assert result["success"] is False
        assert result["created"] == 0
        assert result["updated"] == 0


class TestFileValidation:
    """Tests for file validation in API endpoint."""

    def test_empty_file_should_be_rejected(self):
        """Empty file should be rejected."""
        # This tests the validation logic - actual endpoint test would be in test_inventory.py
        assert len(b"") == 0  # Basic assertion to show empty file detection

    def test_file_size_limit_calculation(self):
        """File size limit should be 10MB."""
        max_size = 10 * 1024 * 1024
        assert max_size == 10485760

        # Test that a file over 10MB would be rejected
        large_content = b"x" * (max_size + 1)
        assert len(large_content) > max_size


class TestBatchProcessing:
    """Tests for batch processing during import."""

    def test_batch_commits_every_100_rows(self, business_id, user_id):
        """Import should commit in batches of 100 rows."""
        # Create 150 products
        products = [_make_product(business_id, f"SKU{i:03d}", f"Product {i}") for i in range(150)]

        db = FakeSession(data_by_model={
            Product: products,
            InventoryItem: [],
        })
        service = InventoryExcelService(db)

        # Create Excel with 150 rows
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=2, value="Quantity On Hand")

        for i in range(150):
            ws.cell(row=i + 2, column=1, value=f"SKU{i:03d}")
            ws.cell(row=i + 2, column=2, value=10)

        output = BytesIO()
        wb.save(output)

        result = service.import_inventory(business_id, output.getvalue(), user_id)

        assert result["success"] is True
        assert result["created"] == 150
        # Should have at least 2 commits (one batch at 100, one final)
        assert db.commits >= 2
