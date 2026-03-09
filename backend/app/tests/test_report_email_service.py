"""Unit tests for ReportEmailService."""

import os

os.environ.setdefault("SECRET_KEY", "test-secret-key")

import io
import logging
from datetime import datetime
from unittest.mock import MagicMock, patch

import pytest

from app.services.email_service import EmailAttachment
from app.services.report_email_service import ReportEmailService


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_report_type(value: str = "daily_sales"):
    """Create a mock ReportType enum member."""
    rt = MagicMock()
    rt.value = value
    return rt


def _make_report_data(*, metrics: dict | None = None, report_type_value: str = "daily_sales"):
    """Build a mock ReportData with sensible defaults."""
    rd = MagicMock()
    rd.report_type = _make_report_type(report_type_value)
    rd.business_name = "Acme Corp"
    rd.business_id = "biz-123"
    rd.user_email = "owner@acme.com"
    rd.period_start = datetime(2024, 1, 1)
    rd.period_end = datetime(2024, 1, 31)
    rd.metrics = metrics if metrics is not None else {
        "total_sales": 15000.50,
        "total_orders": 120,
        "currency": "ZAR",
        "top_products": [
            {"name": "Widget A", "quantity": 50, "revenue": 5000},
            {"name": "Widget B", "quantity": 30, "revenue": 3000},
        ],
        "low_stock_items": [
            {"name": "Gadget X", "sku": "GX-001", "quantity": 2, "reorder_point": 10},
        ],
        "out_of_stock_items": [
            {"name": "Gadget Y", "sku": "GY-002", "quantity": 0, "reorder_point": 5},
        ],
        "top_customers": [
            {"name": "Customer 1", "total_spent": 8000},
        ],
    }
    return rd


@pytest.fixture
def email_service():
    return MagicMock()


@pytest.fixture
def service(email_service):
    return ReportEmailService(email_service)


@pytest.fixture
def report_data():
    return _make_report_data()


# ---------------------------------------------------------------------------
# generate_subject
# ---------------------------------------------------------------------------

class TestGenerateSubject:
    def test_correct_format(self, service, report_data):
        subject = service.generate_subject(report_data)
        assert subject == "Acme Corp: Daily Sales (2024-01-01 to 2024-01-31)"

    def test_different_report_type(self, service):
        rd = _make_report_data(report_type_value="weekly_inventory")
        subject = service.generate_subject(rd)
        assert "Weekly Inventory" in subject

    def test_dates_included(self, service, report_data):
        subject = service.generate_subject(report_data)
        assert "2024-01-01" in subject
        assert "2024-01-31" in subject


# ---------------------------------------------------------------------------
# generate_html_body
# ---------------------------------------------------------------------------

class TestGenerateHtmlBody:
    def test_contains_business_name(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "Acme Corp" in html

    def test_contains_report_name(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "Daily Sales" in html

    def test_contains_period_dates(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "2024-01-01" in html
        assert "2024-01-31" in html

    def test_includes_simple_metrics(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "Total Sales" in html
        assert "15000.5" in html
        assert "Total Orders" in html
        assert "120" in html

    def test_skips_complex_metrics(self, service, report_data):
        """Lists and dicts should not appear in the summary <ul>."""
        html = service.generate_html_body(report_data)
        # The raw key "top_products" should NOT be in the <ul> metrics
        # (it should only appear in its own table heading).
        ul_section = html.split("Summary Metrics")[1].split("</ul>")[0]
        assert "Top Products" not in ul_section

    def test_top_products_table(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "<h3>Top Products</h3>" in html
        assert "Widget A" in html
        assert "Widget B" in html
        assert "5000" in html

    def test_low_stock_table(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "<h3>Low Stock Items</h3>" in html
        assert "Gadget X" in html
        assert "GX-001" in html

    def test_no_low_stock_table_when_empty(self, service):
        rd = _make_report_data(metrics={"total_sales": 100, "low_stock_items": []})
        html = service.generate_html_body(rd)
        assert "Low Stock Items" not in html

    def test_currency_in_product_rows(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "ZAR" in html

    def test_footer_links(self, service, report_data):
        html = service.generate_html_body(report_data)
        assert "/reports/biz-123" in html
        assert "unsubscribe=daily_sales" in html


# ---------------------------------------------------------------------------
# generate_excel_attachment
# ---------------------------------------------------------------------------

class TestGenerateExcelAttachment:
    def test_returns_email_attachment(self, service, report_data):
        att = service.generate_excel_attachment(report_data)
        assert isinstance(att, EmailAttachment)

    def test_correct_filename(self, service, report_data):
        att = service.generate_excel_attachment(report_data)
        assert att.filename == "daily_sales_20240131.xlsx"

    def test_correct_content_type(self, service, report_data):
        att = service.generate_excel_attachment(report_data)
        assert att.content_type == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_content_is_nonempty_bytes(self, service, report_data):
        att = service.generate_excel_attachment(report_data)
        assert isinstance(att.content, bytes)
        assert len(att.content) > 0

    def test_excel_has_summary_sheet(self, service, report_data):
        import openpyxl

        att = service.generate_excel_attachment(report_data)
        wb = openpyxl.load_workbook(io.BytesIO(att.content))
        assert "Summary" in wb.sheetnames

    def test_excel_has_additional_sheets(self, service, report_data):
        import openpyxl

        att = service.generate_excel_attachment(report_data)
        wb = openpyxl.load_workbook(io.BytesIO(att.content))
        assert "Top Products" in wb.sheetnames
        assert "Low Stock" in wb.sheetnames
        assert "Out of Stock" in wb.sheetnames
        assert "Top Customers" in wb.sheetnames

    def test_summary_sheet_excludes_complex_metrics(self, service, report_data):
        import openpyxl

        att = service.generate_excel_attachment(report_data)
        wb = openpyxl.load_workbook(io.BytesIO(att.content))
        ws = wb["Summary"]
        headers = [cell.value for cell in ws[1]]
        assert "total_sales" in headers
        assert "top_products" not in headers

    def test_no_extra_sheets_when_lists_absent(self, service):
        rd = _make_report_data(metrics={"total_sales": 100})
        att = service.generate_excel_attachment(rd)
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(att.content))
        assert wb.sheetnames == ["Summary"]


# ---------------------------------------------------------------------------
# send_report_email
# ---------------------------------------------------------------------------

class TestSendReportEmail:
    def test_calls_send_email(self, service, email_service, report_data):
        service.send_report_email(report_data)
        email_service.send_email.assert_called_once()

    def test_send_email_args(self, service, email_service, report_data):
        service.send_report_email(report_data)
        call_kwargs = email_service.send_email.call_args
        assert call_kwargs.kwargs["to_email"] == "owner@acme.com"
        assert "Acme Corp" in call_kwargs.kwargs["subject"]
        assert call_kwargs.kwargs["attachments"] is None

    def test_include_excel_sends_attachment(self, service, email_service, report_data):
        service.send_report_email(report_data, include_excel=True)
        call_kwargs = email_service.send_email.call_args.kwargs
        assert call_kwargs["attachments"] is not None
        assert len(call_kwargs["attachments"]) == 1
        assert isinstance(call_kwargs["attachments"][0], EmailAttachment)

    def test_excel_error_still_sends_email(self, service, email_service, report_data, caplog):
        with patch.object(
            service, "generate_excel_attachment", side_effect=Exception("pandas boom")
        ):
            with caplog.at_level(logging.ERROR):
                service.send_report_email(report_data, include_excel=True)

        email_service.send_email.assert_called_once()
        call_kwargs = email_service.send_email.call_args.kwargs
        assert call_kwargs["attachments"] is None
        assert "Failed to generate Excel attachment" in caplog.text

    def test_without_excel_no_attachment(self, service, email_service, report_data):
        service.send_report_email(report_data, include_excel=False)
        call_kwargs = email_service.send_email.call_args.kwargs
        assert call_kwargs["attachments"] is None
