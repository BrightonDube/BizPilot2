"""Reports API endpoints."""

from typing import List, Optional
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from sqlalchemy import func

from app.core.database import get_sync_db
from app.api.deps import get_current_active_user, get_current_business_id, check_feature
from app.models.user import User
from app.services.sales_report_service import SalesReportService
from app.services.staff_report_service import StaffReportService
from app.services.inventory_report_service import InventoryReportService
from app.models.business_user import BusinessUser
from app.models.order import Order, OrderDirection
from app.models.customer import Customer
from app.models.product import Product
from app.models.order import OrderItem
from app.models.inventory import InventoryItem
from app.models.session import Session as SessionModel
from app.models.time_entry import TimeEntry
from app.core.pdf import build_simple_pdf
from app.schemas.report import (
    ReportStats,
    TopProduct,
    TopCustomer,
    RevenueTrend,
    OrdersTrend,
    TrendDataPoint,
    InventoryReport,
    InventoryReportItem,
    COGSReport,
    COGSReportItem,
    ProfitMarginReport,
    ProfitMarginItem,
    UserActivityReport,
    UserActivityItem,
    LoginHistoryReport,
    LoginHistoryItem,
)
from app.schemas.staff_report import (
    StaffPerformanceReport,
    StaffAttendanceReport,
    DepartmentPerformanceReport,
    StaffProductivityReport,
    StaffCommissionReport,
    StaffActivityLogReport,
    CashDrawerReport,
    VoidRefundReport,
    DiscountReport,
)
from app.schemas.commission import (
    CommissionGenerateRequest,
    CommissionApproveRequest,
    CommissionListResponse,
    CommissionApproveResponse,
)
from app.services.commission_service import CommissionService
from app.schemas.custom_report import (
    ReportTemplateCreate,
    ReportTemplateUpdate,
    ReportTemplateResponse,
    ReportTemplateListResponse,
    CustomReportRequest,
    CustomReportResponse,
)
from app.services.custom_report_service import CustomReportService

router = APIRouter(prefix="/reports", tags=["Reports"])


def get_date_range(range_str: str) -> tuple[date, date]:
    """Convert range string to start and end dates."""
    today = date.today()
    if range_str == "7d":
        return today - timedelta(days=7), today
    elif range_str == "30d":
        return today - timedelta(days=30), today
    elif range_str == "90d":
        return today - timedelta(days=90), today
    elif range_str == "1y":
        return today - timedelta(days=365), today
    else:
        return today - timedelta(days=30), today


@router.get("/stats", response_model=ReportStats)
async def get_report_stats(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get overall business statistics."""
    
    if not business_id:
        return ReportStats(
            total_revenue=0,
            total_orders=0,
            total_customers=0,
            total_products=0,
            revenue_change=0,
            orders_change=0,
            customers_change=0,
        )

    start_date, end_date = get_date_range(range)
    prev_start = start_date - (end_date - start_date)
    
    # Current period stats
    current_revenue_query = db.query(func.sum(Order.total)).filter(
        Order.business_id == business_id,
        Order.created_at >= start_date,
        Order.created_at <= end_date,
    )
    if direction:
        current_revenue_query = current_revenue_query.filter(Order.direction == direction)
    current_revenue = current_revenue_query.scalar() or 0
    
    current_orders_query = db.query(func.count(Order.id)).filter(
        Order.business_id == business_id,
        Order.created_at >= start_date,
        Order.created_at <= end_date,
    )
    if direction:
        current_orders_query = current_orders_query.filter(Order.direction == direction)
    current_orders = current_orders_query.scalar() or 0
    
    # Previous period stats for comparison
    prev_revenue_query = db.query(func.sum(Order.total)).filter(
        Order.business_id == business_id,
        Order.created_at >= prev_start,
        Order.created_at < start_date,
    )
    if direction:
        prev_revenue_query = prev_revenue_query.filter(Order.direction == direction)
    prev_revenue = prev_revenue_query.scalar() or 0
    
    prev_orders_query = db.query(func.count(Order.id)).filter(
        Order.business_id == business_id,
        Order.created_at >= prev_start,
        Order.created_at < start_date,
    )
    if direction:
        prev_orders_query = prev_orders_query.filter(Order.direction == direction)
    prev_orders = prev_orders_query.scalar() or 0
    
    # Total counts
    total_customers = 0
    if direction in (None, OrderDirection.INBOUND):
        total_customers = db.query(func.count(Customer.id)).filter(
            Customer.business_id == business_id
        ).scalar() or 0
    
    total_products = db.query(func.count(Product.id)).filter(
        Product.business_id == business_id
    ).scalar() or 0
    
    # Calculate percentage changes
    revenue_change = 0
    if prev_revenue > 0:
        revenue_change = round(((float(current_revenue) - float(prev_revenue)) / float(prev_revenue)) * 100, 1)
    
    orders_change = 0
    if prev_orders > 0:
        orders_change = round(((current_orders - prev_orders) / prev_orders) * 100, 1)

    return ReportStats(
        total_revenue=float(current_revenue),
        total_orders=current_orders,
        total_customers=total_customers,
        total_products=total_products,
        revenue_change=revenue_change,
        orders_change=orders_change,
        customers_change=0,
    )


@router.get("/top-products", response_model=List[TopProduct])
async def get_top_products(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    limit: int = Query(5, ge=1, le=20),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get top-selling products."""
    
    if not business_id:
        return []

    # Default to OUTBOUND (sales) for "top products".
    # Both INBOUND and OUTBOUND are valid — the frontend sends INBOUND
    # for the "sales" category and OUTBOUND for "purchases".
    if direction is None:
        direction = OrderDirection.OUTBOUND

    start_date, end_date = get_date_range(range)

    rows = (
        db.query(
            OrderItem.product_id.label("product_id"),
            OrderItem.name.label("name"),
            func.coalesce(func.sum(OrderItem.quantity), 0).label("sales"),
            func.coalesce(func.sum(OrderItem.total), 0).label("revenue"),
        )
        .join(Order, Order.id == OrderItem.order_id)
        .filter(
            Order.business_id == business_id,
            Order.direction == direction,
            Order.created_at >= start_date,
            Order.created_at <= end_date,
            Order.deleted_at.is_(None),
            OrderItem.deleted_at.is_(None),
        )
        .group_by(OrderItem.product_id, OrderItem.name)
        .order_by(func.coalesce(func.sum(OrderItem.total), 0).desc())
        .limit(limit)
        .all()
    )

    return [
        TopProduct(
            id=str(r.product_id) if r.product_id else "",
            name=r.name,
            sales=int(r.sales or 0),
            revenue=float(r.revenue or 0),
        )
        for r in rows
    ]


@router.get("/top-customers", response_model=List[TopCustomer])
async def get_top_customers(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    limit: int = Query(5, ge=1, le=20),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get top customers by spending."""
    
    if not business_id:
        return []

    # Default to OUTBOUND (sales) for "top customers".
    # Both INBOUND and OUTBOUND are valid — the frontend sends INBOUND
    # for the "sales" category and OUTBOUND for "purchases".
    if direction is None:
        direction = OrderDirection.OUTBOUND

    start_date, end_date = get_date_range(range)

    rows = (
        db.query(
            Customer.id.label("customer_id"),
            Customer.first_name.label("first_name"),
            Customer.last_name.label("last_name"),
            Customer.company_name.label("company_name"),
            func.count(Order.id).label("orders"),
            func.coalesce(func.sum(Order.total), 0).label("total_spent"),
        )
        .join(Order, Order.customer_id == Customer.id)
        .filter(
            Order.business_id == business_id,
            Customer.business_id == business_id,
            Order.direction == direction,
            Order.created_at >= start_date,
            Order.created_at <= end_date,
            Order.deleted_at.is_(None),
            Customer.deleted_at.is_(None),
        )
        .group_by(Customer.id, Customer.first_name, Customer.last_name, Customer.company_name)
        .order_by(func.coalesce(func.sum(Order.total), 0).desc())
        .limit(limit)
        .all()
    )

    result: list[TopCustomer] = []
    for r in rows:
        if r.company_name:
            name = r.company_name
        else:
            name = f"{r.first_name or ''} {r.last_name or ''}".strip() or "Unknown"

        result.append(
            TopCustomer(
                id=str(r.customer_id),
                name=name,
                orders=int(r.orders or 0),
                total_spent=float(r.total_spent or 0),
            )
        )

    return result


@router.get("/revenue-trend", response_model=RevenueTrend)
async def get_revenue_trend(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get revenue over time for chart visualization."""
    
    if not business_id:
        return RevenueTrend(data=[], total=0, average=0)

    start_date, end_date = get_date_range(range)
    
    # Determine grouping based on range
    if range == "7d":
        # Group by day
        date_trunc = func.date(Order.created_at)
    elif range == "30d":
        # Group by day
        date_trunc = func.date(Order.created_at)
    elif range == "90d":
        # Group by week
        date_trunc = func.date(Order.created_at)
    else:  # 1y
        # Group by month
        date_trunc = func.date(Order.created_at)
    
    query = (
        db.query(
            date_trunc.label("period"),
            func.coalesce(func.sum(Order.total), 0).label("revenue"),
        )
        .filter(
            Order.business_id == business_id,
            Order.created_at >= start_date,
            Order.created_at <= end_date,
            Order.deleted_at.is_(None),
        )
    )
    
    if direction:
        query = query.filter(Order.direction == direction)
    
    rows = (
        query
        .group_by(date_trunc)
        .order_by(date_trunc)
        .all()
    )
    
    data = []
    total = 0.0
    for r in rows:
        period_str = str(r.period) if r.period else ""
        revenue = float(r.revenue or 0)
        total += revenue
        
        # Format label based on range
        if r.period:
            try:
                from datetime import datetime
                dt = datetime.strptime(period_str, "%Y-%m-%d")
                if range == "7d":
                    label = dt.strftime("%a")  # Mon, Tue, etc.
                elif range == "30d":
                    label = dt.strftime("%d %b")  # 01 Jan
                elif range == "90d":
                    label = dt.strftime("%d %b")  # 01 Jan
                else:
                    label = dt.strftime("%b %Y")  # Jan 2024
            except Exception:
                label = period_str
        else:
            label = period_str
        
        data.append(TrendDataPoint(
            date=period_str,
            value=revenue,
            label=label,
        ))
    
    average = total / len(data) if data else 0
    
    return RevenueTrend(data=data, total=total, average=average)


@router.get("/orders-trend", response_model=OrdersTrend)
async def get_orders_trend(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get orders count over time for chart visualization."""
    
    if not business_id:
        return OrdersTrend(data=[], total=0, average=0)

    start_date, end_date = get_date_range(range)
    
    date_trunc = func.date(Order.created_at)
    
    query = (
        db.query(
            date_trunc.label("period"),
            func.count(Order.id).label("order_count"),
        )
        .filter(
            Order.business_id == business_id,
            Order.created_at >= start_date,
            Order.created_at <= end_date,
            Order.deleted_at.is_(None),
        )
    )
    
    if direction:
        query = query.filter(Order.direction == direction)
    
    rows = (
        query
        .group_by(date_trunc)
        .order_by(date_trunc)
        .all()
    )
    
    data = []
    total = 0
    for r in rows:
        period_str = str(r.period) if r.period else ""
        count = int(r.order_count or 0)
        total += count
        
        # Format label based on range
        if r.period:
            try:
                from datetime import datetime
                dt = datetime.strptime(period_str, "%Y-%m-%d")
                if range == "7d":
                    label = dt.strftime("%a")  # Mon, Tue, etc.
                elif range == "30d":
                    label = dt.strftime("%d %b")  # 01 Jan
                elif range == "90d":
                    label = dt.strftime("%d %b")  # 01 Jan
                else:
                    label = dt.strftime("%b %Y")  # Jan 2024
            except Exception:
                label = period_str
        else:
            label = period_str
        
        data.append(TrendDataPoint(
            date=period_str,
            value=float(count),
            label=label,
        ))
    
    average = total / len(data) if data else 0
    
    return OrdersTrend(data=data, total=total, average=average)


@router.get("/inventory", response_model=InventoryReport)
async def get_inventory_report(
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get inventory status report."""
    
    if not business_id:
        return InventoryReport(
            items=[],
            total_items=0,
            total_value=0,
            low_stock_count=0,
            out_of_stock_count=0,
        )

    rows = (
        db.query(InventoryItem)
        .join(Product, Product.id == InventoryItem.product_id)
        .filter(
            InventoryItem.business_id == business_id,
            InventoryItem.deleted_at.is_(None),
            Product.deleted_at.is_(None),
        )
        .all()
    )

    items = []
    total_value = 0.0
    low_stock_count = 0
    out_of_stock_count = 0

    for item in rows:
        quantity = float(item.quantity_on_hand or 0)
        reorder_point = float(item.reorder_point or 0)
        unit_cost = float(item.unit_cost or 0)
        item_value = quantity * unit_cost
        total_value += item_value

        if quantity <= 0:
            status = "out_of_stock"
            out_of_stock_count += 1
        elif quantity <= reorder_point:
            status = "low_stock"
            low_stock_count += 1
        else:
            status = "in_stock"

        items.append(InventoryReportItem(
            id=str(item.id),
            product_name=item.product_name or "Unknown",
            sku=item.sku,
            quantity_on_hand=quantity,
            reorder_point=reorder_point,
            unit_cost=unit_cost,
            total_value=item_value,
            status=status,
        ))

    return InventoryReport(
        items=items,
        total_items=len(items),
        total_value=total_value,
        low_stock_count=low_stock_count,
        out_of_stock_count=out_of_stock_count,
    )


@router.get("/cogs", response_model=COGSReport)
async def get_cogs_report(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get Cost of Goods Sold report."""
    
    if not business_id:
        return COGSReport(
            items=[],
            total_cogs=0,
            total_revenue=0,
            gross_profit=0,
            average_margin=0,
        )

    start_date, end_date = get_date_range(range)

    # Get sales data with product costs
    rows = (
        db.query(
            OrderItem.product_id.label("product_id"),
            OrderItem.name.label("product_name"),
            func.sum(OrderItem.quantity).label("quantity_sold"),
            func.sum(OrderItem.total).label("revenue"),
        )
        .join(Order, Order.id == OrderItem.order_id)
        .filter(
            Order.business_id == business_id,
            Order.direction == OrderDirection.INBOUND,
            Order.created_at >= start_date,
            Order.created_at <= end_date,
            Order.deleted_at.is_(None),
            OrderItem.deleted_at.is_(None),
        )
        .group_by(OrderItem.product_id, OrderItem.name)
        .all()
    )

    items = []
    total_cogs = 0.0
    total_revenue = 0.0

    for r in rows:
        product_id = str(r.product_id) if r.product_id else ""
        quantity_sold = int(r.quantity_sold or 0)
        revenue = float(r.revenue or 0)
        total_revenue += revenue

        # Get product cost
        product = db.query(Product).filter(Product.id == r.product_id).first()
        unit_cost = float(product.total_cost or 0) if product else 0
        total_cost = unit_cost * quantity_sold
        total_cogs += total_cost

        gross_profit = revenue - total_cost
        margin_percent = (gross_profit / revenue * 100) if revenue > 0 else 0

        items.append(COGSReportItem(
            product_id=product_id,
            product_name=r.product_name or "Unknown",
            quantity_sold=quantity_sold,
            unit_cost=unit_cost,
            total_cost=total_cost,
            revenue=revenue,
            gross_profit=gross_profit,
            margin_percent=round(margin_percent, 1),
        ))

    gross_profit = total_revenue - total_cogs
    average_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

    return COGSReport(
        items=items,
        total_cogs=total_cogs,
        total_revenue=total_revenue,
        gross_profit=gross_profit,
        average_margin=round(average_margin, 1),
    )


@router.get("/profit-margins", response_model=ProfitMarginReport)
async def get_profit_margins_report(
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get profit margins by product."""
    
    if not business_id:
        return ProfitMarginReport(
            items=[],
            average_margin=0,
            highest_margin=0,
            lowest_margin=0,
        )

    products = (
        db.query(Product)
        .filter(
            Product.business_id == business_id,
            Product.deleted_at.is_(None),
        )
        .all()
    )

    items = []
    margins = []

    for product in products:
        selling_price = float(product.selling_price or 0)
        total_cost = float(product.total_cost or 0)
        profit = selling_price - total_cost
        margin_percent = (profit / selling_price * 100) if selling_price > 0 else 0

        items.append(ProfitMarginItem(
            product_id=str(product.id),
            product_name=product.name,
            selling_price=selling_price,
            total_cost=total_cost,
            profit=profit,
            margin_percent=round(margin_percent, 1),
        ))
        margins.append(margin_percent)

    average_margin = sum(margins) / len(margins) if margins else 0
    highest_margin = max(margins) if margins else 0
    lowest_margin = min(margins) if margins else 0

    # Sort by margin descending
    items.sort(key=lambda x: x.margin_percent, reverse=True)

    return ProfitMarginReport(
        items=items,
        average_margin=round(average_margin, 1),
        highest_margin=round(highest_margin, 1),
        lowest_margin=round(lowest_margin, 1),
    )


@router.get("/export/pdf")
async def export_reports_pdf(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    direction: Optional[OrderDirection] = Query(None),
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    start_date, end_date = get_date_range(range)

    lines: list[str] = []
    lines.append("BizPilot Report")
    lines.append(f"Range: {start_date} to {end_date}")
    if direction:
        lines.append(f"Category: {direction}")
    lines.append("")

    if not business_id:
        lines.append("No business selected.")
        pdf_bytes = build_simple_pdf(lines)
        filename = f"report_{range}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    stats = await get_report_stats(range=range, direction=direction, current_user=current_user, db=db)
    top_products = await get_top_products(range=range, limit=5, direction=direction, current_user=current_user, db=db)
    top_customers = await get_top_customers(range=range, limit=5, direction=direction, current_user=current_user, db=db)

    lines.append("Summary")
    lines.append(f"Total Revenue: R {stats.total_revenue}")
    lines.append(f"Total Orders: {stats.total_orders}")
    lines.append(f"Total Customers: {stats.total_customers}")
    lines.append(f"Total Products: {stats.total_products}")
    lines.append("")
    lines.append("Top Products")

    if not top_products:
        lines.append("(No product data)")
    else:
        for idx, p in enumerate(top_products, start=1):
            lines.append(f"{idx}. {p.name} | Sales: {p.sales} | Revenue: R {p.revenue}")

    lines.append("")
    lines.append("Top Customers")

    if not top_customers:
        lines.append("(No customer data)")
    else:
        for idx, c in enumerate(top_customers, start=1):
            lines.append(f"{idx}. {c.name} | Orders: {c.orders} | Total Spent: R {c.total_spent}")

    pdf_bytes = build_simple_pdf(lines)
    direction_str = f"_{direction}" if direction else ""
    filename = f"report{direction_str}_{range}_{start_date}_to_{end_date}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



@router.get("/user-activity", response_model=UserActivityReport)
async def get_user_activity_report(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    user_id: Optional[str] = Query(None),
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get user activity report with time tracking data."""
    
    if not business_id:
        return UserActivityReport(
            items=[],
            total_users=0,
            total_hours=0,
            average_hours_per_user=0,
        )

    start_date, end_date = get_date_range(range)
    
    # Build query for time entries
    query = (
        db.query(
            TimeEntry.user_id,
            User.first_name,
            User.last_name,
            func.sum(TimeEntry.hours_worked).label("total_hours"),
            func.count(TimeEntry.id).label("total_entries"),
            func.sum(
                func.case(
                    (TimeEntry.clock_in.isnot(None), 1),
                    else_=0
                )
            ).label("clock_ins"),
            func.sum(
                func.case(
                    (TimeEntry.clock_out.isnot(None), 1),
                    else_=0
                )
            ).label("clock_outs"),
            func.sum(TimeEntry.break_duration).label("break_duration"),
            func.max(TimeEntry.clock_in).label("last_activity"),
        )
        .join(User, User.id == TimeEntry.user_id)
        .filter(
            TimeEntry.business_id == business_id,
            TimeEntry.clock_in >= start_date,
            TimeEntry.clock_in <= end_date,
            TimeEntry.deleted_at.is_(None),
        )
    )
    
    if user_id:
        query = query.filter(TimeEntry.user_id == user_id)
    
    rows = (
        query
        .group_by(TimeEntry.user_id, User.first_name, User.last_name)
        .order_by(func.sum(TimeEntry.hours_worked).desc())
        .all()
    )
    
    items = []
    total_hours = 0.0
    
    for r in rows:
        user_name = f"{r.first_name or ''} {r.last_name or ''}".strip() or "Unknown"
        hours = float(r.total_hours or 0)
        total_hours += hours
        
        # Check if user has any active sessions (clock_in without clock_out)
        has_active = db.query(TimeEntry).filter(
            TimeEntry.user_id == r.user_id,
            TimeEntry.business_id == business_id,
            TimeEntry.clock_in.isnot(None),
            TimeEntry.clock_out.is_(None),
            TimeEntry.deleted_at.is_(None),
        ).first() is not None
        
        items.append(UserActivityItem(
            user_id=str(r.user_id),
            user_name=user_name,
            total_hours=round(hours, 2),
            total_entries=int(r.total_entries or 0),
            clock_ins=int(r.clock_ins or 0),
            clock_outs=int(r.clock_outs or 0),
            break_duration=round(float(r.break_duration or 0), 2),
            last_activity=r.last_activity.isoformat() if r.last_activity else None,
            status="active" if has_active else "completed",
        ))
    
    average_hours = total_hours / len(items) if items else 0
    
    return UserActivityReport(
        items=items,
        total_users=len(items),
        total_hours=round(total_hours, 2),
        average_hours_per_user=round(average_hours, 2),
    )


@router.get("/login-history", response_model=LoginHistoryReport)
async def get_login_history_report(
    range: str = Query("30d", pattern="^(7d|30d|90d|1y)$"),
    user_id: Optional[str] = Query(None),
    include_active: bool = Query(True),
    current_user: User = Depends(check_feature("has_advanced_reporting")),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get login history report with session tracking data."""
    
    if not business_id:
        return LoginHistoryReport(
            items=[],
            total_sessions=0,
            active_sessions=0,
            unique_users=0,
            suspicious_count=0,
        )

    start_date, end_date = get_date_range(range)
    
    # Build query for sessions
    # Need to join through business_user to filter by business
    query = (
        db.query(SessionModel, User)
        .join(User, User.id == SessionModel.user_id)
        .join(BusinessUser, BusinessUser.user_id == User.id)
        .filter(
            BusinessUser.business_id == business_id,
            SessionModel.created_at >= start_date,
            SessionModel.created_at <= end_date,
        )
    )
    
    if user_id:
        query = query.filter(SessionModel.user_id == user_id)
    
    if not include_active:
        query = query.filter(~SessionModel.is_active)
    
    rows = query.order_by(SessionModel.created_at.desc()).all()
    
    items = []
    active_count = 0
    suspicious_count = 0
    unique_users = set()
    
    # Track IPs per user for suspicious activity detection
    user_ips = {}
    
    for session, user in rows:
        unique_users.add(str(session.user_id))
        
        # Calculate duration
        duration_minutes = None
        logout_time = None
        
        if session.revoked_at:
            logout_time = session.revoked_at.isoformat()
            duration = session.revoked_at - session.created_at
            duration_minutes = round(duration.total_seconds() / 60, 1)
        elif not session.is_active:
            # Session expired
            logout_time = session.expires_at.isoformat()
            duration = session.expires_at - session.created_at
            duration_minutes = round(duration.total_seconds() / 60, 1)
        
        if session.is_active:
            active_count += 1
        
        # Detect suspicious activity
        is_suspicious = False
        user_id_str = str(session.user_id)
        
        # Track IPs for this user
        if user_id_str not in user_ips:
            user_ips[user_id_str] = set()
        
        if session.ip_address:
            user_ips[user_id_str].add(session.ip_address)
            
            # Flag if user has multiple IPs in this period (possible account sharing)
            if len(user_ips[user_id_str]) > 3:
                is_suspicious = True
        
        # Flag very long sessions (>24 hours)
        if duration_minutes and duration_minutes > 1440:
            is_suspicious = True
        
        if is_suspicious:
            suspicious_count += 1
        
        user_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.email or "Unknown"
        
        items.append(LoginHistoryItem(
            session_id=str(session.id),
            user_id=str(session.user_id),
            user_name=user_name,
            device_name=session.device_name,
            device_type=session.device_type,
            ip_address=session.ip_address,
            location=session.location,
            login_time=session.created_at.isoformat(),
            logout_time=logout_time,
            duration_minutes=duration_minutes,
            is_active=session.is_active,
            is_suspicious=is_suspicious,
        ))
    
    return LoginHistoryReport(
        items=items,
        total_sessions=len(items),
        active_sessions=active_count,
        unique_users=len(unique_users),
        suspicious_count=suspicious_count,
    )


# ---------------------------------------------------------------------------
# Sales Report Endpoints
# ---------------------------------------------------------------------------


@router.get("/sales/daily")
async def get_daily_sales_report(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get daily sales report with hourly breakdown."""
    from datetime import date as date_type

    if date:
        try:
            target_date = date_type.fromisoformat(date)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD.",
            )
    else:
        target_date = date_type.today()

    service = SalesReportService(db)
    return service.get_daily_report(business_id, target_date)


@router.get("/sales/weekly")
async def get_weekly_sales_report(
    week_start: Optional[str] = Query(None, description="Monday date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get weekly sales report with daily breakdown."""
    from datetime import date as date_type

    if week_start:
        try:
            start = date_type.fromisoformat(week_start)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD.",
            )
    else:
        today = date_type.today()
        start = today - timedelta(days=today.weekday())  # Monday

    service = SalesReportService(db)
    return service.get_weekly_report(business_id, start)


@router.get("/sales/monthly")
async def get_monthly_sales_report(
    year: Optional[int] = Query(None, description="Year (e.g. 2024)"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Month (1-12)"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get monthly sales report with daily breakdown."""
    from datetime import date as date_type

    today = date_type.today()
    year = year or today.year
    month = month or today.month

    service = SalesReportService(db)
    return service.get_monthly_report(business_id, year, month)


@router.get("/sales/products")
async def get_product_performance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get product performance report ranked by revenue."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    return service.get_product_performance(business_id, start, end, limit)


@router.get("/sales/categories")
async def get_category_performance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get category performance report."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    return service.get_category_performance(business_id, start, end)


@router.get("/sales/payments")
async def get_payment_breakdown_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get payment method breakdown report."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    try:
        return service.get_payment_breakdown(business_id, start, end)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating payment breakdown report: {str(e)}",
        )


@router.get("/sales/time-analysis")
async def get_time_analysis_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get time-based analysis with peak hours and day-of-week breakdown."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    return service.get_time_analysis(business_id, start, end)


@router.get("/sales/discounts")
async def get_discount_analysis_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get discount analysis report with breakdown by product and category."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    return service.get_discount_analysis(business_id, start, end)


@router.get("/sales/refunds")
async def get_refund_analysis_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get refund analysis report with breakdown by product and trend."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    return service.get_refund_analysis(business_id, start, end)


@router.get("/export/csv")
async def export_report_csv(
    report_type: str = Query(
        ...,
        description="Report type: products, categories, payments, discounts, refunds",
    ),
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export sales report data as CSV."""
    import csv
    import io
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    output = io.StringIO()
    writer = csv.writer(output)

    if report_type == "products":
        data = service.get_product_performance(business_id, start, end)
        writer.writerow(["Rank", "Product", "Quantity Sold", "Revenue", "Order Count", "Revenue %"])
        for p in data.get("products", []):
            writer.writerow([
                p["rank"], p["product_name"], p["quantity_sold"],
                p["revenue"], p["order_count"], p["revenue_percentage"],
            ])
    elif report_type == "categories":
        data = service.get_category_performance(business_id, start, end)
        writer.writerow(["Category", "Quantity Sold", "Revenue", "Order Count", "Revenue %"])
        for c in data.get("categories", []):
            writer.writerow([
                c["category_name"], c["quantity_sold"],
                c["revenue"], c["order_count"], c["revenue_percentage"],
            ])
    elif report_type == "payments":
        data = service.get_payment_breakdown(business_id, start, end)
        writer.writerow(["Payment Method", "Count", "Amount", "Amount %", "Count %"])
        for m in data.get("methods", []):
            writer.writerow([
                m["payment_method"], m["count"], m["amount"],
                m["percentage_amount"], m["percentage_count"],
            ])
    elif report_type == "discounts":
        data = service.get_discount_analysis(business_id, start, end)
        writer.writerow(["Product", "Discount Total", "Revenue", "Discount %", "Item Count"])
        for p in data.get("by_product", []):
            writer.writerow([
                p["product_name"], p["discount_total"], p["revenue"],
                p["discount_percentage"], p["item_count"],
            ])
    elif report_type == "refunds":
        data = service.get_refund_analysis(business_id, start, end)
        writer.writerow(["Product", "Refund Total", "Quantity", "Refund Count", "% of Refunds"])
        for p in data.get("by_product", []):
            writer.writerow([
                p["product_name"], p["refund_total"], p["quantity"],
                p["refund_count"], p["percentage_of_refunds"],
            ])
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown report type: {report_type}. Use: products, categories, payments, discounts, refunds",
        )

    csv_content = output.getvalue()
    filename = f"{report_type}_report_{start_date}_to_{end_date}.csv"
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/excel")
async def export_report_excel(
    report_type: str = Query(
        ...,
        description="Report type: products, categories, payments, discounts, refunds",
    ),
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export sales report data as Excel (.xlsx) with styled headers."""
    from io import BytesIO
    from datetime import date as date_type

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = SalesReportService(db)
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    def write_headers(headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_row(row_num: int, values: list) -> None:
        is_alt = row_num % 2 == 0
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

    if report_type == "products":
        ws.title = "Product Performance"
        data = service.get_product_performance(business_id, start, end)
        write_headers(["Rank", "Product", "Quantity Sold", "Revenue", "Order Count", "Revenue %"])
        for i, p in enumerate(data.get("products", []), 2):
            write_row(i, [
                p["rank"], p["product_name"], p["quantity_sold"],
                p["revenue"], p["order_count"], p["revenue_percentage"],
            ])
    elif report_type == "categories":
        ws.title = "Category Performance"
        data = service.get_category_performance(business_id, start, end)
        write_headers(["Category", "Quantity Sold", "Revenue", "Order Count", "Revenue %"])
        for i, c in enumerate(data.get("categories", []), 2):
            write_row(i, [
                c["category_name"], c["quantity_sold"],
                c["revenue"], c["order_count"], c["revenue_percentage"],
            ])
    elif report_type == "payments":
        ws.title = "Payment Breakdown"
        data = service.get_payment_breakdown(business_id, start, end)
        write_headers(["Payment Method", "Count", "Amount", "Amount %", "Count %"])
        for i, m in enumerate(data.get("methods", []), 2):
            write_row(i, [
                m["payment_method"], m["count"], m["amount"],
                m["percentage_amount"], m["percentage_count"],
            ])
    elif report_type == "discounts":
        ws.title = "Discount Analysis"
        data = service.get_discount_analysis(business_id, start, end)
        write_headers(["Product", "Discount Total", "Revenue", "Discount %", "Item Count"])
        for i, p in enumerate(data.get("by_product", []), 2):
            write_row(i, [
                p["product_name"], p["discount_total"], p["revenue"],
                p["discount_percentage"], p["item_count"],
            ])
    elif report_type == "refunds":
        ws.title = "Refund Analysis"
        data = service.get_refund_analysis(business_id, start, end)
        write_headers(["Product", "Refund Total", "Quantity", "Refund Count", "% of Refunds"])
        for i, p in enumerate(data.get("by_product", []), 2):
            write_row(i, [
                p["product_name"], p["refund_total"], p["quantity"],
                p["refund_count"], p["percentage_of_refunds"],
            ])
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown report type: {report_type}. Use: products, categories, payments, discounts, refunds",
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{report_type}_report_{start_date}_to_{end_date}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Staff Reports ──────────────────────────────────────────────────────────


@router.get("/staff/performance", response_model=StaffPerformanceReport)
async def get_staff_performance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Staff performance report with hours worked and ranking."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_performance_report(business_id, start, end)


@router.get("/staff/attendance", response_model=StaffAttendanceReport)
async def get_staff_attendance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    user_id: Optional[str] = Query(None, description="Filter by specific user ID"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Staff attendance report with daily breakdown and summaries."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_attendance_report(business_id, start, end, user_id=user_id)


@router.get("/staff/departments", response_model=DepartmentPerformanceReport)
async def get_department_performance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Department performance report with staff and hours breakdown."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_department_performance(business_id, start, end)


@router.get("/staff/productivity", response_model=StaffProductivityReport)
async def get_staff_productivity_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Staff productivity analysis with efficiency metrics."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_productivity_report(business_id, start, end)


@router.get("/staff/commissions", response_model=StaffCommissionReport)
async def get_staff_commission_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    commission_rate: float = Query(5.0, description="Commission rate percentage"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Commission report based on sales attributed to staff."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_commission_report(business_id, start, end, commission_rate)


@router.get("/staff/activity", response_model=StaffActivityLogReport)
async def get_staff_activity_log(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    action_type: Optional[str] = Query(None, description="Filter by action type"),
    resource_type: Optional[str] = Query(None, description="Filter by resource type"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(50, ge=1, le=100, description="Items per page"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Activity log report from audit trail."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_activity_log(
        business_id, start, end,
        user_id=user_id,
        action_type=action_type,
        resource_type=resource_type,
        page=page,
        per_page=per_page,
    )


@router.get("/staff/cash-drawer", response_model=CashDrawerReport)
async def get_cash_drawer_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    register_id: Optional[str] = Query(None, description="Filter by register ID"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Cash drawer report with register sessions and cash movements."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_cash_drawer_report(
        business_id, start, end,
        register_id=register_id,
        user_id=user_id,
    )


# ── Staff Report Exports ────────────────────────────────────────────────


STAFF_REPORT_TYPES = [
    "performance", "attendance", "department", "productivity",
    "commission", "voids-refunds", "discounts", "cash-drawer",
]


@router.get("/staff/export/csv")
async def export_staff_report_csv(
    report_type: str = Query(..., description="Staff report type"),
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    commission_rate: float = Query(5.0, description="Commission rate %"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export staff report as CSV."""
    import csv
    import io
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if report_type not in STAFF_REPORT_TYPES:
        raise HTTPException(status_code=400, detail=f"Unknown type: {report_type}. Use: {', '.join(STAFF_REPORT_TYPES)}")

    service = StaffReportService(db)
    output = io.StringIO()
    writer = csv.writer(output)

    if report_type == "performance":
        data = service.get_performance_report(business_id, start, end)
        writer.writerow(["Rank", "Staff", "Email", "Total Hours", "Net Hours", "Sales Count", "Sales Total"])
        for s in data.get("staff", []):
            writer.writerow([s["rank"], s["staff_name"], s["email"], s["total_hours"], s["net_hours"], s["sales_count"], s["sales_total"]])
    elif report_type == "attendance":
        data = service.get_attendance_report(business_id, start, end)
        writer.writerow(["Staff", "Email", "Days Present", "Days Absent", "Late Count", "Total Hours", "Attendance Rate"])
        for s in data.get("staff", []):
            writer.writerow([s["staff_name"], s["email"], s["days_present"], s["days_absent"], s["late_count"], s["total_hours"], s["attendance_rate"]])
    elif report_type == "commission":
        data = service.get_commission_report(business_id, start, end, commission_rate)
        writer.writerow(["Rank", "Staff", "Email", "Orders", "Total Sales", "Discounts", "Commission Rate", "Commission Amount"])
        for s in data.get("staff", []):
            writer.writerow([s["rank"], s["staff_name"], s["email"], s["order_count"], s["total_sales"], s["total_discounts"], s["commission_rate"], s["commission_amount"]])
    elif report_type == "voids-refunds":
        data = service.get_void_refund_report(business_id, start, end)
        writer.writerow(["Staff", "Email", "Voids", "Refunds", "Total Actions"])
        for s in data.get("staff", []):
            writer.writerow([s["staff_name"], s["email"], s["void_count"], s["refund_count"], s["total_actions"]])
    elif report_type == "discounts":
        data = service.get_discount_report(business_id, start, end)
        writer.writerow(["Staff", "Email", "Orders", "Total Discounts", "Total Sales", "Discount Rate %"])
        for s in data.get("staff", []):
            writer.writerow([s["staff_name"], s["email"], s["order_count"], s["total_discounts"], s["total_sales"], s["discount_rate"]])
    else:
        raise HTTPException(status_code=400, detail=f"CSV export not supported for: {report_type}")

    csv_content = output.getvalue()
    filename = f"staff_{report_type}_{start_date}_to_{end_date}.csv"
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/staff/export/excel")
async def export_staff_report_excel(
    report_type: str = Query(..., description="Staff report type"),
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    commission_rate: float = Query(5.0, description="Commission rate %"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export staff report as styled Excel (.xlsx)."""
    from io import BytesIO
    from datetime import date as date_type
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if report_type not in STAFF_REPORT_TYPES:
        raise HTTPException(status_code=400, detail=f"Unknown type: {report_type}. Use: {', '.join(STAFF_REPORT_TYPES)}")

    service = StaffReportService(db)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Staff {report_type.replace('-', ' ').title()}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if report_type == "performance":
        data = service.get_performance_report(business_id, start, end)
        headers = ["Rank", "Staff", "Email", "Total Hours", "Net Hours", "Sales Count", "Sales Total"]
        rows = [[s["rank"], s["staff_name"], s["email"], s["total_hours"], s["net_hours"], s["sales_count"], s["sales_total"]] for s in data.get("staff", [])]
    elif report_type == "attendance":
        data = service.get_attendance_report(business_id, start, end)
        headers = ["Staff", "Email", "Days Present", "Days Absent", "Late Count", "Total Hours", "Attendance Rate"]
        rows = [[s["staff_name"], s["email"], s["days_present"], s["days_absent"], s["late_count"], s["total_hours"], s["attendance_rate"]] for s in data.get("staff", [])]
    elif report_type == "commission":
        data = service.get_commission_report(business_id, start, end, commission_rate)
        headers = ["Rank", "Staff", "Email", "Orders", "Total Sales", "Discounts", "Rate", "Commission"]
        rows = [[s["rank"], s["staff_name"], s["email"], s["order_count"], s["total_sales"], s["total_discounts"], s["commission_rate"], s["commission_amount"]] for s in data.get("staff", [])]
    elif report_type == "voids-refunds":
        data = service.get_void_refund_report(business_id, start, end)
        headers = ["Staff", "Email", "Voids", "Refunds", "Total Actions"]
        rows = [[s["staff_name"], s["email"], s["void_count"], s["refund_count"], s["total_actions"]] for s in data.get("staff", [])]
    elif report_type == "discounts":
        data = service.get_discount_report(business_id, start, end)
        headers = ["Staff", "Email", "Orders", "Total Discounts", "Total Sales", "Discount Rate %"]
        rows = [[s["staff_name"], s["email"], s["order_count"], s["total_discounts"], s["total_sales"], s["discount_rate"]] for s in data.get("staff", [])]
    else:
        raise HTTPException(status_code=400, detail=f"Excel export not supported for: {report_type}")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"staff_{report_type}_{start_date}_to_{end_date}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/staff/export/pdf")
async def export_staff_report_pdf(
    report_type: str = Query(..., description="Staff report type"),
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    commission_rate: float = Query(5.0, description="Commission rate %"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export staff report as PDF."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if report_type not in STAFF_REPORT_TYPES:
        raise HTTPException(status_code=400, detail=f"Unknown type: {report_type}. Use: {', '.join(STAFF_REPORT_TYPES)}")

    service = StaffReportService(db)
    title = f"Staff {report_type.replace('-', ' ').title()} Report"
    subtitle = f"{start_date} to {end_date}"

    if report_type == "performance":
        data = service.get_performance_report(business_id, start, end)
        headers = ["Rank", "Staff", "Hours", "Sales", "Total"]
        rows = [[s["rank"], s["staff_name"], s["total_hours"], s["sales_count"], s["sales_total"]] for s in data.get("staff", [])]
    elif report_type == "attendance":
        data = service.get_attendance_report(business_id, start, end)
        headers = ["Staff", "Present", "Absent", "Late", "Rate"]
        rows = [[s["staff_name"], s["days_present"], s["days_absent"], s["late_count"], s["attendance_rate"]] for s in data.get("staff", [])]
    elif report_type == "commission":
        data = service.get_commission_report(business_id, start, end, commission_rate)
        headers = ["Staff", "Orders", "Sales", "Commission"]
        rows = [[s["staff_name"], s["order_count"], s["total_sales"], s["commission_amount"]] for s in data.get("staff", [])]
    elif report_type == "voids-refunds":
        data = service.get_void_refund_report(business_id, start, end)
        headers = ["Staff", "Voids", "Refunds", "Total"]
        rows = [[s["staff_name"], s["void_count"], s["refund_count"], s["total_actions"]] for s in data.get("staff", [])]
    elif report_type == "discounts":
        data = service.get_discount_report(business_id, start, end)
        headers = ["Staff", "Orders", "Discounts", "Rate %"]
        rows = [[s["staff_name"], s["order_count"], s["total_discounts"], s["discount_rate"]] for s in data.get("staff", [])]
    else:
        raise HTTPException(status_code=400, detail=f"PDF export not supported for: {report_type}")

    pdf_bytes = build_simple_pdf(title, subtitle, headers, rows)
    filename = f"staff_{report_type}_{start_date}_to_{end_date}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Void/Discount Reports ───────────────────────────────────────────────


@router.get("/staff/voids-refunds", response_model=VoidRefundReport)
async def get_void_refund_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Void and refund report showing actions by staff."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_void_refund_report(business_id, start, end, user_id=user_id)


@router.get("/staff/discounts", response_model=DiscountReport)
async def get_discount_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Discount report showing discounts applied by staff."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = StaffReportService(db)
    return service.get_discount_report(business_id, start, end, user_id=user_id)


# ── Commission Approval Workflow ────────────────────────────────────────


@router.post("/staff/commissions/generate", response_model=CommissionListResponse)
async def generate_commission_records(
    body: CommissionGenerateRequest,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Generate commission records from sales data for a period."""
    report_service = StaffReportService(db)
    report = report_service.get_commission_report(
        business_id, body.period_start, body.period_end, body.commission_rate,
    )
    commission_service = CommissionService(db)
    commission_service.generate_records(
        business_id=business_id,
        period_start=body.period_start,
        period_end=body.period_end,
        commission_rate=body.commission_rate,
        staff_data=report["staff"],
    )
    items, total = commission_service.list_records(business_id)
    total_commission = sum(i["commission_amount"] for i in items)
    total_sales = sum(i["total_sales"] for i in items)
    return CommissionListResponse(
        items=items, total=total,
        total_commission=total_commission, total_sales=total_sales,
    )


@router.get("/staff/commissions/records", response_model=CommissionListResponse)
async def list_commission_records(
    status_filter: Optional[str] = Query(None, alias="status", description="Filter by status"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """List commission records with optional status filter."""
    commission_service = CommissionService(db)
    items, total = commission_service.list_records(
        business_id, status=status_filter, page=page, per_page=per_page,
    )
    total_commission = sum(i["commission_amount"] for i in items)
    total_sales = sum(i["total_sales"] for i in items)
    return CommissionListResponse(
        items=items, total=total,
        total_commission=total_commission, total_sales=total_sales,
    )


@router.post("/staff/commissions/approve", response_model=CommissionApproveResponse)
async def approve_commission_records(
    body: CommissionApproveRequest,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Approve or reject commission records."""
    commission_service = CommissionService(db)
    if body.action == "approve":
        count = commission_service.approve_records(
            business_id, body.record_ids, current_user.id,
        )
    else:
        count = commission_service.reject_records(
            business_id, body.record_ids, current_user.id,
            reason=body.rejection_reason,
        )
    return CommissionApproveResponse(updated_count=count, action=body.action)


@router.get("/staff/commissions/payroll")
async def export_commission_payroll(
    period_start: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    period_end: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Export approved commissions grouped by staff for payroll."""
    from datetime import date as date_type

    start = date_type.fromisoformat(period_start) if period_start else None
    end = date_type.fromisoformat(period_end) if period_end else None

    commission_service = CommissionService(db)
    return commission_service.get_payroll_export(business_id, start, end)


# ── Custom Report Builder ───────────────────────────────────────────────


@router.post("/custom/execute", response_model=CustomReportResponse)
async def execute_custom_report(
    body: CustomReportRequest,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Execute a custom report with selected metrics and filters."""
    service = CustomReportService(db)
    return service.execute_report(
        business_id, body.metrics, body.filters,
        body.group_by, body.sort_by, body.sort_direction,
    )


@router.post("/custom/templates", response_model=ReportTemplateResponse)
async def create_report_template(
    body: ReportTemplateCreate,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Save a custom report template."""
    service = CustomReportService(db)
    template = service.create_template(
        business_id=business_id, user_id=current_user.id,
        name=body.name, description=body.description,
        metrics=body.metrics, filters=body.filters,
        group_by=body.group_by, sort_by=body.sort_by,
        sort_direction=body.sort_direction,
        is_scheduled=body.is_scheduled, schedule_cron=body.schedule_cron,
        schedule_recipients=body.schedule_recipients,
        is_public=body.is_public,
    )
    owner_name = f"{current_user.first_name or ''} {current_user.last_name or ''}".strip() or "Unknown"
    return ReportTemplateResponse(
        id=template.id, name=template.name, description=template.description,
        report_type=template.report_type, metrics=template.metrics,
        filters=template.filters, group_by=template.group_by,
        sort_by=template.sort_by, sort_direction=template.sort_direction,
        is_scheduled=template.is_scheduled, schedule_cron=template.schedule_cron,
        schedule_recipients=template.schedule_recipients,
        is_public=template.is_public, created_by_name=owner_name,
        created_at=template.created_at,
    )


@router.get("/custom/templates", response_model=ReportTemplateListResponse)
async def list_report_templates(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """List saved report templates."""
    service = CustomReportService(db)
    items, total = service.list_templates(
        business_id, user_id=current_user.id, page=page, per_page=per_page,
    )
    return ReportTemplateListResponse(items=items, total=total)


@router.get("/custom/templates/{template_id}", response_model=ReportTemplateResponse)
async def get_report_template(
    template_id: str,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get a specific report template."""
    service = CustomReportService(db)
    template = service.get_template(business_id, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    owner_name = "Unknown"
    if template.owner:
        owner_name = f"{template.owner.first_name or ''} {template.owner.last_name or ''}".strip() or "Unknown"
    return ReportTemplateResponse(
        id=template.id, name=template.name, description=template.description,
        report_type=template.report_type, metrics=template.metrics,
        filters=template.filters, group_by=template.group_by,
        sort_by=template.sort_by, sort_direction=template.sort_direction,
        is_scheduled=template.is_scheduled, schedule_cron=template.schedule_cron,
        schedule_recipients=template.schedule_recipients,
        is_public=template.is_public, created_by_name=owner_name,
        created_at=template.created_at,
    )


@router.put("/custom/templates/{template_id}", response_model=ReportTemplateResponse)
async def update_report_template(
    template_id: str,
    body: ReportTemplateUpdate,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Update a report template."""
    service = CustomReportService(db)
    template = service.get_template(business_id, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    if template.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this template")
    updates = body.model_dump(exclude_unset=True)
    template = service.update_template(template, **updates)
    owner_name = f"{current_user.first_name or ''} {current_user.last_name or ''}".strip() or "Unknown"
    return ReportTemplateResponse(
        id=template.id, name=template.name, description=template.description,
        report_type=template.report_type, metrics=template.metrics,
        filters=template.filters, group_by=template.group_by,
        sort_by=template.sort_by, sort_direction=template.sort_direction,
        is_scheduled=template.is_scheduled, schedule_cron=template.schedule_cron,
        schedule_recipients=template.schedule_recipients,
        is_public=template.is_public, created_by_name=owner_name,
        created_at=template.created_at,
    )


@router.delete("/custom/templates/{template_id}")
async def delete_report_template(
    template_id: str,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Delete a report template."""
    service = CustomReportService(db)
    template = service.get_template(business_id, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    if template.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this template")
    service.delete_template(template)
    return {"detail": "Template deleted"}


@router.get("/custom/metrics")
async def list_available_metrics(
    current_user: User = Depends(get_current_active_user),
):
    """List available metrics for custom report builder."""
    from app.services.custom_report_service import AVAILABLE_METRICS, AVAILABLE_GROUP_BY
    return {
        "metrics": [{"key": k, "label": v} for k, v in AVAILABLE_METRICS.items()],
        "group_by": AVAILABLE_GROUP_BY,
    }


@router.get("/inventory/stock-levels")
async def get_stock_levels_report(
    category_id: Optional[str] = Query(None, description="Filter by category ID"),
    low_stock_only: bool = Query(False, description="Show only low stock items"),
    out_of_stock_only: bool = Query(False, description="Show only out of stock items"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get current stock levels with low/out-of-stock flags."""
    service = InventoryReportService(db)
    return service.get_stock_levels(
        business_id,
        category_id=category_id,
        low_stock_only=low_stock_only,
        out_of_stock_only=out_of_stock_only,
    )


@router.get("/inventory/movements")
async def get_stock_movements_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    product_id: Optional[str] = Query(None, description="Filter by product ID"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get stock movement report showing ins, outs, and adjustments."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = InventoryReportService(db)
    return service.get_stock_movements(business_id, start, end, product_id=product_id)


@router.get("/inventory/valuation")
async def get_inventory_valuation_report(
    method: str = Query("average", description="Valuation method (average)"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get inventory valuation report grouped by category."""
    service = InventoryReportService(db)
    return service.get_valuation(business_id, method=method)


@router.get("/inventory/turnover")
async def get_inventory_turnover_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get inventory turnover analysis with fast/slow/dead stock classification."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = InventoryReportService(db)
    return service.get_turnover_analysis(business_id, start, end)


@router.get("/inventory/supplier-performance")
async def get_supplier_performance_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get supplier performance report based on purchase orders."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = InventoryReportService(db)
    return service.get_supplier_performance(business_id, start, end)


@router.get("/inventory/wastage")
async def get_wastage_report(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get inventory wastage/shrinkage report."""
    from datetime import date as date_type

    try:
        start = date_type.fromisoformat(start_date)
        end = date_type.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    service = InventoryReportService(db)
    return service.get_wastage_report(business_id, start, end)


@router.get("/inventory/dashboard")
async def get_inventory_dashboard(
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Get inventory dashboard with KPIs, alerts, and stock overview."""
    service = InventoryReportService(db)
    return service.get_inventory_dashboard(business_id)


# ── Modifier & Combo Reports ──────────────────────────────────────
# Tasks 9.1-9.4 from addons-modifiers spec (Requirement 10)

from app.services.modifier_analytics_service import ModifierAnalyticsService


@router.get("/modifiers/summary")
async def get_modifier_summary(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """High-level modifier statistics for the dashboard."""
    service = ModifierAnalyticsService(db)
    return service.get_modifier_summary(business_id, start_date, end_date)


@router.get("/modifiers/frequency")
async def get_modifier_frequency(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    product_id: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Modifier selection frequency, ordered by popularity."""
    try:
        service = ModifierAnalyticsService(db)
        return service.get_modifier_selection_frequency(
            business_id, start_date, end_date, product_id=product_id, limit=limit,
        )
    except Exception:
        return []


@router.get("/modifiers/revenue")
async def get_modifier_revenue(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Revenue generated by each modifier."""
    try:
        service = ModifierAnalyticsService(db)
        return service.get_modifier_revenue(business_id, start_date, end_date, limit=limit)
    except Exception:
        return []


@router.get("/modifiers/rankings")
async def get_modifier_rankings(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    order_by: str = Query("popular", description="popular, unpopular, revenue, aov_impact"),
    limit: int = Query(10, ge=1, le=50),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Ranked modifier list by popularity or revenue."""
    try:
        service = ModifierAnalyticsService(db)
        return service.get_modifier_rankings(
            business_id, start_date, end_date, order_by=order_by, limit=limit,
        )
    except Exception:
        return []


@router.get("/combos/performance")
async def get_combo_performance(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db=Depends(get_sync_db),
):
    """Combo deal performance metrics."""
    service = ModifierAnalyticsService(db)
    return service.get_combo_performance(business_id, start_date, end_date, limit=limit)
