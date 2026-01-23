"""Time Entry API endpoints for time tracking and payroll."""

import io
import math
from typing import Optional, Dict, List
from datetime import datetime, date, timedelta
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, ConfigDict

# Import openpyxl at module level for efficiency
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.core.database import get_db
from app.api.deps import get_current_active_user, get_current_business_id, check_feature
from app.core.rbac import has_permission
from app.models.user import User
from app.models.time_entry import TimeEntry, TimeEntryStatus
from app.services.time_tracking_service import TimeTrackingService

router = APIRouter(prefix="/time-entries", tags=["Time Tracking"])


# --- Schemas ---

class TimeEntryCreate(BaseModel):
    """Schema for creating a time entry via clock in."""
    device_id: Optional[str] = None
    ip_address: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None


class TimeEntryManualCreate(BaseModel):
    """Schema for creating a manual time entry."""
    user_id: str
    clock_in: datetime
    clock_out: datetime
    notes: Optional[str] = None


class TimeEntryUpdate(BaseModel):
    """Schema for updating a time entry."""
    notes: Optional[str] = None


class TimeEntryApproval(BaseModel):
    """Schema for approving/rejecting a time entry."""
    approved: bool
    rejection_reason: Optional[str] = None


class TimeEntryResponse(BaseModel):
    """Response schema for a time entry."""
    model_config = ConfigDict(from_attributes=True)
    
    id: str
    user_id: str
    user_name: Optional[str] = None
    business_id: str
    entry_type: str
    clock_in: Optional[datetime] = None
    clock_out: Optional[datetime] = None
    break_start: Optional[datetime] = None
    break_end: Optional[datetime] = None
    hours_worked: Optional[float] = None
    break_duration: Optional[float] = None
    status: str
    device_id: Optional[str] = None
    ip_address: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    approved_by_id: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class TimeEntryListResponse(BaseModel):
    """Response schema for a list of time entries."""
    items: list[TimeEntryResponse]
    total: int
    page: int
    per_page: int
    pages: int


class UserTimeSummary(BaseModel):
    """Summary of user's time for a period."""
    user_id: str
    date_from: str
    date_to: str
    total_hours: float
    total_break_hours: float
    days_worked: int
    entries_count: int
    average_hours_per_day: float


class PayrollReportItem(BaseModel):
    """Individual user's payroll data."""
    user_id: str
    user_name: str
    email: str
    total_hours: float
    total_break_hours: float
    entries_count: int


class PayrollReport(BaseModel):
    """Payroll report for all users."""
    date_from: str
    date_to: str
    items: list[PayrollReportItem]
    total_hours: float


class ClockStatus(BaseModel):
    """Current clock status for a user."""
    is_clocked_in: bool
    entry_id: Optional[str] = None
    clock_in: Optional[datetime] = None
    is_on_break: bool = False
    break_start: Optional[datetime] = None


class TimeEntryInlineUpdate(BaseModel):
    """Schema for inline editing time entries."""
    clock_in: Optional[datetime] = None
    clock_out: Optional[datetime] = None
    break_start: Optional[datetime] = None
    break_end: Optional[datetime] = None
    break_duration: Optional[float] = None
    notes: Optional[str] = None


class CurrentlyWorkingUser(BaseModel):
    """Schema for currently working user."""
    user_id: str
    name: str
    email: str
    clock_in: str
    current_hours: float
    on_break: bool
    break_start: Optional[str] = None


class DayEndProcessResult(BaseModel):
    """Schema for day-end process result."""
    auto_clocked_out: int
    penalty_hours: Optional[float] = None
    message: str


class BusinessTimeSettingsResponse(BaseModel):
    """Schema for business time settings."""
    model_config = ConfigDict(from_attributes=True)
    
    id: str
    business_id: str
    day_end_time: str
    auto_clock_out_penalty_hours: float
    standard_work_hours: float
    overtime_threshold: float
    payroll_period_type: str
    payroll_period_start_day: int


class BusinessTimeSettingsUpdate(BaseModel):
    """Schema for updating business time settings."""
    day_end_time: Optional[str] = None  # Format: "HH:MM"
    auto_clock_out_penalty_hours: Optional[float] = None
    standard_work_hours: Optional[float] = None
    overtime_threshold: Optional[float] = None
    payroll_period_type: Optional[str] = None
    payroll_period_start_day: Optional[int] = None


def _entry_to_response(entry: TimeEntry, user: Optional[User] = None) -> TimeEntryResponse:
    """Convert time entry to response."""
    return TimeEntryResponse(
        id=str(entry.id),
        user_id=str(entry.user_id),
        user_name=user.full_name if user else None,
        business_id=str(entry.business_id),
        entry_type=entry.entry_type.value if entry.entry_type else None,
        clock_in=entry.clock_in,
        clock_out=entry.clock_out,
        break_start=entry.break_start,
        break_end=entry.break_end,
        hours_worked=float(entry.hours_worked) if entry.hours_worked else None,
        break_duration=float(entry.break_duration) if entry.break_duration else None,
        status=entry.status.value if entry.status else None,
        device_id=entry.device_id,
        ip_address=entry.ip_address,
        location=entry.location,
        notes=entry.notes,
        approved_by_id=str(entry.approved_by_id) if entry.approved_by_id else None,
        approved_at=entry.approved_at,
        rejection_reason=entry.rejection_reason,
        created_at=entry.created_at,
        updated_at=entry.updated_at,
    )


# --- Endpoints ---

@router.get("/status", response_model=ClockStatus)
async def get_clock_status(
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get current clock status for the authenticated user."""
    service = TimeTrackingService(db)
    entry = service._get_active_entry(business_id, str(current_user.id))
    
    if entry:
        is_on_break = entry.break_start is not None and entry.break_end is None
        return ClockStatus(
            is_clocked_in=True,
            entry_id=str(entry.id),
            clock_in=entry.clock_in,
            is_on_break=is_on_break,
            break_start=entry.break_start if is_on_break else None,
        )
    
    return ClockStatus(is_clocked_in=False)


@router.post("/clock-in", response_model=TimeEntryResponse)
async def clock_in(
    data: TimeEntryCreate,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Clock in for the current user."""
    service = TimeTrackingService(db)
    
    try:
        entry = service.clock_in(
            business_id=business_id,
            user_id=str(current_user.id),
            device_id=data.device_id,
            location=data.location,
        )
        if data.notes:
            entry.notes = data.notes
            db.commit()
            db.refresh(entry)
        return _entry_to_response(entry, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/clock-out", response_model=TimeEntryResponse)
async def clock_out(
    data: TimeEntryUpdate,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Clock out for the current user."""
    service = TimeTrackingService(db)
    
    try:
        entry = service.clock_out(
            business_id=business_id,
            user_id=str(current_user.id),
        )
        if data.notes:
            entry.notes = data.notes
            db.commit()
            db.refresh(entry)
        return _entry_to_response(entry, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/break/start", response_model=TimeEntryResponse)
async def start_break(
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Start a break for the current user."""
    service = TimeTrackingService(db)
    
    try:
        entry = service.start_break(business_id, str(current_user.id))
        return _entry_to_response(entry, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/break/end", response_model=TimeEntryResponse)
async def end_break(
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """End a break for the current user."""
    service = TimeTrackingService(db)
    
    try:
        entry = service.end_break(business_id, str(current_user.id))
        return _entry_to_response(entry, current_user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("", response_model=TimeEntryListResponse)
async def list_time_entries(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    user_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """List time entries with filtering."""
    service = TimeTrackingService(db)
    
    # Parse status filter
    entry_status = None
    if status_filter:
        try:
            entry_status = TimeEntryStatus(status_filter)
        except ValueError:
            # Invalid status filter provided; ignore and treat as no status filter
            pass
    
    # Use the new service method to get entries for a user or team
    if user_id:
        entries = service.get_user_time_entries(business_id, user_id, date_from, date_to)
    else:
        # Get team entries and flatten
        team_entries = service.get_team_time_entries(business_id, date_from, date_to)
        entries = []
        for user_entries in team_entries.values():
            entries.extend(user_entries)
        # Sort by clock_in descending
        entries.sort(key=lambda x: x.clock_in or datetime.min, reverse=True)
    
    # Filter by status if provided
    if entry_status:
        entries = [e for e in entries if e.status == entry_status]
    
    # Apply pagination
    total = len(entries)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_entries = entries[start_idx:end_idx]
    
    # Get users for names
    user_map = {}
    for entry in paginated_entries:
        if entry.user_id not in user_map:
            user = db.query(User).filter(User.id == entry.user_id).first()
            user_map[entry.user_id] = user
    
    return TimeEntryListResponse(
        items=[_entry_to_response(e, user_map.get(e.user_id)) for e in paginated_entries],
        total=total,
        page=page,
        per_page=per_page,
        pages=math.ceil(total / per_page) if total > 0 else 0,
    )


@router.get("/summary/me", response_model=UserTimeSummary)
async def get_my_time_summary(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(get_current_active_user),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get time summary for the current user."""
    service = TimeTrackingService(db)
    
    # Default to current week
    if not date_to:
        date_to = date.today()
    if not date_from:
        date_from = date_to - timedelta(days=7)
    
    entries = service.get_user_time_entries(business_id, str(current_user.id), date_from, date_to)
    
    # Calculate summary
    completed_entries = [e for e in entries if e.clock_out]
    total_hours = sum(e.net_hours or Decimal("0") for e in completed_entries)
    total_breaks = sum(e.break_duration or Decimal("0") for e in completed_entries)
    days_worked = len(set(e.clock_in.date() for e in completed_entries if e.clock_in))
    
    return UserTimeSummary(
        user_id=str(current_user.id),
        date_from=date_from.isoformat(),
        date_to=date_to.isoformat(),
        total_hours=float(total_hours),
        total_break_hours=float(total_breaks),
        days_worked=days_worked,
        entries_count=len(completed_entries),
        average_hours_per_day=float(total_hours / days_worked) if days_worked > 0 else 0,
    )


@router.get("/summary/{user_id}", response_model=UserTimeSummary)
async def get_user_time_summary(
    user_id: str,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(has_permission("reports:view")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get time summary for a specific user (requires reports:view permission)."""
    service = TimeTrackingService(db)
    
    # Default to current week
    if not date_to:
        date_to = date.today()
    if not date_from:
        date_from = date_to - timedelta(days=7)
    
    entries = service.get_user_time_entries(business_id, user_id, date_from, date_to)
    
    # Calculate summary
    completed_entries = [e for e in entries if e.clock_out]
    total_hours = sum(e.net_hours or Decimal("0") for e in completed_entries)
    total_breaks = sum(e.break_duration or Decimal("0") for e in completed_entries)
    days_worked = len(set(e.clock_in.date() for e in completed_entries if e.clock_in))
    
    return UserTimeSummary(
        user_id=user_id,
        date_from=date_from.isoformat(),
        date_to=date_to.isoformat(),
        total_hours=float(total_hours),
        total_break_hours=float(total_breaks),
        days_worked=days_worked,
        entries_count=len(completed_entries),
        average_hours_per_day=float(total_hours / days_worked) if days_worked > 0 else 0,
    )


@router.get("/payroll-report", response_model=PayrollReport, dependencies=[Depends(check_feature("has_payroll"))])
async def get_payroll_report(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(has_permission("reports:view")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get payroll report for all users (requires reports:view permission and payroll feature)."""
    service = TimeTrackingService(db)
    
    # Default to current pay period (2 weeks)
    if not date_to:
        date_to = date.today()
    if not date_from:
        date_from = date_to - timedelta(days=14)
    
    report = service.get_payroll_report(business_id, date_from, date_to)
    
    total_hours = sum(item["total_hours"] for item in report)
    
    return PayrollReport(
        date_from=date_from.isoformat(),
        date_to=date_to.isoformat(),
        items=[PayrollReportItem(
            user_id=item["user_id"],
            user_name=item["employee"],
            email=item["email"],
            total_hours=item["total_hours"],
            total_break_hours=item["break_hours"],
            entries_count=item["entries"]
        ) for item in report],
        total_hours=total_hours,
    )


@router.get("/payroll-report/export", dependencies=[Depends(check_feature("has_payroll"))])
async def export_payroll_report(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(has_permission("reports:view")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Export payroll report to Excel (requires reports:view permission and payroll feature)."""
    # openpyxl imported at module level for efficiency
    
    service = TimeTrackingService(db)
    
    # Default to current pay period
    if not date_to:
        date_to = date.today()
    if not date_from:
        date_from = date_to - timedelta(days=14)
    
    report = service.get_payroll_report(business_id, date_from, date_to)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Payroll Report: {date_from.isoformat()} to {date_to.isoformat()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Employee Name", "Email", "Total Hours", "Break Hours", "Net Hours", "Entries"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, item in enumerate(report, 4):
        net_hours = item["total_hours"] - item["break_hours"]
        row_data = [
            item["employee"],
            item["email"],
            round(item["total_hours"], 2),
            round(item["break_hours"], 2),
            round(net_hours, 2),
            item["entries"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Total row
    total_row = len(report) + 4
    total_hours = sum(item["total_hours"] for item in report)
    total_breaks = sum(item["break_hours"] for item in report)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=round(total_hours, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(total_breaks, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(total_hours - total_breaks, 2)).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"payroll_report_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/manual", response_model=TimeEntryResponse)
async def create_manual_entry(
    data: TimeEntryManualCreate,
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Create a manual time entry (requires users:manage permission)."""
    if data.clock_out < data.clock_in:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Clock out time must be after clock in time."
        )
    
    # Create manual entry directly
    entry = TimeEntry(
        user_id=data.user_id,
        business_id=business_id,
        clock_in=data.clock_in,
        clock_out=data.clock_out,
        status=TimeEntryStatus.PENDING_APPROVAL,
        notes=f"[Manual Entry] {data.notes}" if data.notes else "[Manual Entry]",
    )
    
    # Calculate hours
    total_seconds = (data.clock_out - data.clock_in).total_seconds()
    entry.hours_worked = Decimal(str(total_seconds)) / Decimal("3600")
    entry.net_hours = entry.hours_worked
    
    db.add(entry)
    db.commit()
    db.refresh(entry)
    
    return _entry_to_response(entry)


@router.patch("/{entry_id}/approve", response_model=TimeEntryResponse)
async def approve_time_entry(
    entry_id: str,
    data: TimeEntryApproval,
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Approve or reject a time entry (requires users:manage permission)."""
    entry = db.query(TimeEntry).filter(
        TimeEntry.id == entry_id,
        TimeEntry.business_id == business_id,
        TimeEntry.deleted_at.is_(None),
    ).first()
    
    if not entry:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Time entry not found"
        )
    
    if data.approved:
        entry.status = TimeEntryStatus.APPROVED
        entry.approved_by_id = current_user.id
        entry.approved_at = datetime.now()
    else:
        if not data.rejection_reason:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Rejection reason is required"
            )
        entry.status = TimeEntryStatus.REJECTED
        entry.rejection_reason = data.rejection_reason
    
    db.commit()
    db.refresh(entry)
    
    return _entry_to_response(entry)


# --- New Advanced Endpoints ---

@router.patch("/{entry_id}/inline-edit", response_model=TimeEntryResponse)
async def inline_edit_time_entry(
    entry_id: str,
    data: TimeEntryInlineUpdate,
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Inline edit time entry (admin/manager only)."""
    service = TimeTrackingService(db)
    
    try:
        entry = service.update_time_entry(
            entry_id=entry_id,
            clock_in=data.clock_in,
            clock_out=data.clock_out,
            break_start=data.break_start,
            break_end=data.break_end,
            break_duration=Decimal(str(data.break_duration)) if data.break_duration is not None else None,
            notes=data.notes,
        )
        return _entry_to_response(entry)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/team", response_model=Dict[str, List[TimeEntryResponse]])
async def get_team_time_entries(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(has_permission("reports:view")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get team time entries grouped by user."""
    service = TimeTrackingService(db)
    
    # Default to current month
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date.replace(day=1)
    
    team_entries = service.get_team_time_entries(business_id, start_date, end_date)
    
    # Get user info and convert to response format
    result = {}
    for user_id, entries in team_entries.items():
        user = db.query(User).filter(User.id == user_id).first()
        result[user_id] = [_entry_to_response(e, user) for e in entries]
    
    return result


@router.get("/currently-working", response_model=List[CurrentlyWorkingUser])
async def get_currently_working_users(
    current_user: User = Depends(has_permission("reports:view")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get list of currently working users."""
    service = TimeTrackingService(db)
    working_users = service.get_currently_working_users(business_id)
    
    return [CurrentlyWorkingUser(**user) for user in working_users]


@router.post("/day-end-process", response_model=DayEndProcessResult)
async def run_day_end_process(
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Run day-end process to auto clock-out employees."""
    service = TimeTrackingService(db)
    result = service.run_day_end_process(business_id)
    
    return DayEndProcessResult(**result)


@router.get("/business-settings", response_model=BusinessTimeSettingsResponse, dependencies=[Depends(check_feature("has_payroll"))])
async def get_business_time_settings(
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Get business time tracking settings (requires users:manage permission and payroll feature)."""
    service = TimeTrackingService(db)
    settings = service.get_or_create_business_settings(business_id)
    
    return BusinessTimeSettingsResponse(
        id=str(settings.id),
        business_id=str(settings.business_id),
        day_end_time=settings.day_end_time_str,
        auto_clock_out_penalty_hours=float(settings.auto_clock_out_penalty_hours),
        standard_work_hours=float(settings.standard_work_hours),
        overtime_threshold=float(settings.overtime_threshold),
        payroll_period_type=settings.payroll_period_type,
        payroll_period_start_day=settings.payroll_period_start_day,
    )


@router.patch("/business-settings", response_model=BusinessTimeSettingsResponse, dependencies=[Depends(check_feature("has_payroll"))])
async def update_business_time_settings(
    data: BusinessTimeSettingsUpdate,
    current_user: User = Depends(has_permission("users:manage")),
    business_id: str = Depends(get_current_business_id),
    db: Session = Depends(get_db),
):
    """Update business time tracking settings (requires users:manage permission and payroll feature)."""
    from datetime import time as dt_time
    
    service = TimeTrackingService(db)
    settings = service.get_or_create_business_settings(business_id)
    
    if data.day_end_time:
        try:
            hour, minute = map(int, data.day_end_time.split(':'))
            settings.day_end_time = dt_time(hour, minute)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid time format. Use HH:MM"
            )
    
    if data.auto_clock_out_penalty_hours is not None:
        settings.auto_clock_out_penalty_hours = Decimal(str(data.auto_clock_out_penalty_hours))
    
    if data.standard_work_hours is not None:
        settings.standard_work_hours = Decimal(str(data.standard_work_hours))
    
    if data.overtime_threshold is not None:
        settings.overtime_threshold = Decimal(str(data.overtime_threshold))
    
    if data.payroll_period_type:
        settings.payroll_period_type = data.payroll_period_type
    
    if data.payroll_period_start_day is not None:
        settings.payroll_period_start_day = data.payroll_period_start_day
    
    db.commit()
    db.refresh(settings)
    
    return BusinessTimeSettingsResponse(
        id=str(settings.id),
        business_id=str(settings.business_id),
        day_end_time=settings.day_end_time_str,
        auto_clock_out_penalty_hours=float(settings.auto_clock_out_penalty_hours),
        standard_work_hours=float(settings.standard_work_hours),
        overtime_threshold=float(settings.overtime_threshold),
        payroll_period_type=settings.payroll_period_type,
        payroll_period_start_day=settings.payroll_period_start_day,
    )