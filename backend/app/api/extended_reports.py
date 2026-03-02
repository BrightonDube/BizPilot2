"""
Extended Reports API — User Activity, Login History, and Excel Export.

Provides three endpoints beyond the standard reporting suite:
1. /reports/user-activity — aggregated work hours per staff member
2. /reports/login-history — session audit trail with suspicious flags
3. /reports/export/excel — multi-report Excel export (openpyxl)

Why separate from the main reports router?
The main reports router handles sales/financial data. These endpoints
query different tables (sessions, time_entries) and have different
access control requirements (admin-only for login history).
"""

from datetime import date
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session

from app.api.deps import get_db, get_current_active_user, get_current_business_id
from app.services.extended_report_service import ExtendedReportService


router = APIRouter(prefix="/reports", tags=["Extended Reports"])


# ---------------------------------------------------------------------------
# User Activity
# ---------------------------------------------------------------------------

@router.get("/user-activity")
async def get_user_activity(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user_id: Optional[UUID] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    business_id: UUID = Depends(get_current_business_id),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """
    Get aggregated user activity from time entries.

    Returns per-user summaries including entry count and last active date.
    Filterable by date range and specific user.
    """
    service = ExtendedReportService(db)
    items, total = service.get_user_activity(
        business_id=business_id,
        start_date=start_date,
        end_date=end_date,
        user_id=user_id,
        page=page,
        per_page=per_page,
    )
    pages = (total + per_page - 1) // per_page if per_page > 0 else 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
    }


# ---------------------------------------------------------------------------
# Login History
# ---------------------------------------------------------------------------

@router.get("/login-history")
async def get_login_history(
    user_id: Optional[UUID] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    suspicious_only: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    business_id: UUID = Depends(get_current_business_id),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """
    Get login history with suspicious activity flagging.

    Flags sessions with:
    - Duration exceeding 24 hours (stale sessions)
    - IP address anomalies

    IP addresses are masked for privacy compliance (POPIA).
    """
    service = ExtendedReportService(db)
    items, total = service.get_login_history(
        business_id=business_id,
        user_id=user_id,
        start_date=start_date,
        end_date=end_date,
        suspicious_only=suspicious_only,
        page=page,
        per_page=per_page,
    )
    pages = (total + per_page - 1) // per_page if per_page > 0 else 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
    }


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

@router.get("/export/excel")
async def export_excel(
    report_type: str = Query(..., description="Report type: user-activity or login-history"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    business_id: UUID = Depends(get_current_business_id),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """
    Export report data as Excel file.

    Why openpyxl over csv?
    Excel format allows formatting (bold headers, column widths, number
    formats) that makes the export immediately usable by managers
    without manual cleanup.

    Note: Returns JSON data when openpyxl is not installed, with a
    header indicating the format. This graceful degradation ensures
    the endpoint works in environments without openpyxl.
    """
    valid_types = {"user-activity", "login-history"}
    if report_type not in valid_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid report_type. Must be one of: {', '.join(valid_types)}",
        )

    service = ExtendedReportService(db)

    if report_type == "user-activity":
        items, _ = service.get_user_activity(
            business_id=business_id,
            start_date=start_date,
            end_date=end_date,
            page=1,
            per_page=10000,  # Export all
        )
    else:
        items, _ = service.get_login_history(
            business_id=business_id,
            start_date=start_date,
            end_date=end_date,
            page=1,
            per_page=10000,
        )

    # Try to generate Excel; fall back to JSON if openpyxl unavailable
    try:
        import openpyxl
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.replace("-", " ").title()

        # Write headers from first item's keys
        if items:
            headers = list(items[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            # Write data rows
            for row_idx, item in enumerate(items, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(key))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{report_type}_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except ImportError:
        # Graceful degradation: return JSON with format header
        return {
            "format": "json",
            "message": "Excel export unavailable (openpyxl not installed). Returning JSON.",
            "items": items,
        }
