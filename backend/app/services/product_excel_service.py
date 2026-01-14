"""Excel import/export service for product management."""

from io import BytesIO
from datetime import datetime, timezone
from decimal import Decimal
from typing import Dict, Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.product import Product, ProductStatus
from app.models.inventory import InventoryItem


# Column definitions for product import/export
PRODUCT_COLUMNS = [
    {"key": "sku", "header": "SKU", "width": 15, "required": False},
    {"key": "name", "header": "Product Name", "width": 30, "required": True},
    {"key": "description", "header": "Description", "width": 40, "required": False},
    {"key": "barcode", "header": "Barcode", "width": 18, "required": False},
    {"key": "selling_price", "header": "Selling Price", "width": 15, "required": True},
    {"key": "cost_price", "header": "Cost Price", "width": 15, "required": False},
    {"key": "quantity", "header": "Initial Quantity", "width": 18, "required": False},
    {"key": "low_stock_threshold", "header": "Low Stock Threshold", "width": 20, "required": False},
    {"key": "is_taxable", "header": "Is Taxable (Y/N)", "width": 15, "required": False},
    {"key": "status", "header": "Status", "width": 12, "required": False},
]

# Constants
IMPORT_BATCH_SIZE = 100  # Commit every N rows during import


class ProductExcelService:
    """Service for Excel-based product import/export."""

    def __init__(self, db: Session):
        self.db = db

    def _apply_header_style(self, cell: Cell) -> None:
        """Apply styling to header cells."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _apply_data_style(self, cell: Cell, is_alternate: bool = False) -> None:
        """Apply styling to data cells."""
        if is_alternate:
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_template(self) -> BytesIO:
        """Generate an empty Excel template with correct column headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Add headers
        for col_idx, col_def in enumerate(PRODUCT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Add instructions sheet
        instructions = wb.create_sheet("Instructions")
        instructions["A1"] = "BizPilot Product Import Template"
        instructions["A1"].font = Font(bold=True, size=14)
        
        instructions["A3"] = "Column Descriptions:"
        instructions["A3"].font = Font(bold=True)
        
        row = 4
        for col_def in PRODUCT_COLUMNS:
            required = " (Required)" if col_def["required"] else " (Optional)"
            instructions[f"A{row}"] = f"• {col_def['header']}{required}"
            row += 1

        instructions["A" + str(row + 1)] = "Notes:"
        instructions["A" + str(row + 1)].font = Font(bold=True)
        instructions[f"A{row + 2}"] = "• Product Name and Selling Price are required for each row"
        instructions[f"A{row + 3}"] = "• SKU must be unique if provided"
        instructions[f"A{row + 4}"] = "• Prices should be numeric values without currency symbols"
        instructions[f"A{row + 5}"] = "• Is Taxable: Enter Y or N (defaults to Y if blank)"
        instructions[f"A{row + 6}"] = "• Status: active, draft, or archived (defaults to active)"
        instructions[f"A{row + 7}"] = "• Products will automatically be added to inventory with specified quantity"

        # Freeze header row
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_products(self, business_id: str) -> BytesIO:
        """Export all products to Excel spreadsheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Add headers
        for col_idx, col_def in enumerate(PRODUCT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Get products
        products = (
            self.db.query(Product)
            .filter(
                Product.business_id == business_id,
                Product.deleted_at.is_(None),
            )
            .order_by(Product.name)
            .all()
        )

        # Add data rows
        for row_idx, product in enumerate(products, 2):
            is_alternate = row_idx % 2 == 0
            
            row_data = [
                product.sku,
                product.name,
                product.description,
                product.barcode,
                float(product.selling_price) if product.selling_price else 0,
                float(product.cost_price) if product.cost_price else None,
                product.quantity,
                product.low_stock_threshold,
                "Y" if product.is_taxable else "N",
                product.status.value if product.status else "active",
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_data_style(cell, is_alternate)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add metadata sheet
        meta = wb.create_sheet("Export Info")
        meta["A1"] = "Export Date:"
        meta["B1"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        meta["A2"] = "Total Products:"
        meta["B2"] = len(products)
        meta["A3"] = "Business ID:"
        meta["B3"] = str(business_id)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def import_products(
        self, business_id: str, file_content: bytes, auto_create_inventory: bool = True
    ) -> Dict[str, Any]:
        """
        Import products from Excel spreadsheet.
        
        Args:
            business_id: The business ID to import products for
            file_content: The Excel file content as bytes
            auto_create_inventory: If True, automatically create inventory items for new products
        
        Returns dict with:
        - success: bool
        - updated: int (count of updated products)
        - created: int (count of new products)
        - errors: list of error messages
        - skipped: int (count of skipped rows)
        """
        result: Dict[str, Any] = {
            "success": True,
            "updated": 0,
            "created": 0,
            "errors": [],
            "skipped": 0,
        }

        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        except Exception as e:
            result["success"] = False
            result["errors"].append(f"Failed to read Excel file: {str(e)}")
            return result

        # Find the products sheet
        ws = None
        if not wb.sheetnames:
            result["success"] = False
            result["errors"].append("Excel file contains no sheets")
            return result
            
        for sheet_name in ["Products", "Sheet1"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # Get header row and map columns
        headers: Dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header_lower = str(cell.value).lower().strip().replace(" ", "_")
                headers[header_lower] = col_idx

        # Validate required columns
        required_headers = {"product_name", "selling_price"}
        # Also accept "name" as alias for "product_name"
        if "name" in headers and "product_name" not in headers:
            headers["product_name"] = headers["name"]
        
        missing = required_headers - set(headers.keys())
        if missing:
            result["success"] = False
            result["errors"].append(f"Missing required columns: {', '.join(missing)}")
            return result

        # Pre-fetch all products for this business (for SKU duplicate checking)
        existing_products = (
            self.db.query(Product)
            .filter(
                Product.business_id == business_id,
                Product.deleted_at.is_(None),
            )
            .all()
        )
        products_by_sku: Dict[str, Product] = {p.sku: p for p in existing_products if p.sku}

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                def get_cell_value(header_key: str, default=None):
                    col = headers.get(header_key)
                    if col is None:
                        return default
                    val = row[col - 1].value
                    return val if val is not None else default

                # Get product name (required)
                name = get_cell_value("product_name")
                if not name:
                    result["skipped"] += 1
                    continue
                name = str(name).strip()

                # Get selling price (required)
                selling_price = get_cell_value("selling_price")
                if selling_price is None:
                    result["errors"].append(f"Row {row_idx}: Missing selling price")
                    result["skipped"] += 1
                    continue
                
                try:
                    selling_price = Decimal(str(selling_price))
                except Exception:
                    result["errors"].append(f"Row {row_idx}: Invalid selling price value")
                    result["skipped"] += 1
                    continue

                # Get SKU (optional but must be unique)
                sku = get_cell_value("sku")
                if sku:
                    sku = str(sku).strip()
                    # Check for duplicates
                    if sku in products_by_sku:
                        # Update existing product
                        product = products_by_sku[sku]
                        is_new = False
                    else:
                        product = Product(
                            business_id=UUID(business_id),
                            sku=sku,
                        )
                        is_new = True
                else:
                    # No SKU - always create new
                    product = Product(business_id=UUID(business_id))
                    is_new = True

                # Set required fields
                product.name = name
                product.selling_price = selling_price

                # Optional fields
                description = get_cell_value("description")
                if description is not None:
                    product.description = str(description).strip() or None

                barcode = get_cell_value("barcode")
                if barcode is not None:
                    product.barcode = str(barcode).strip() or None

                cost_price = get_cell_value("cost_price")
                if cost_price is not None:
                    try:
                        product.cost_price = Decimal(str(cost_price))
                    except Exception:
                        pass  # Keep existing or default

                quantity = get_cell_value("initial_quantity") or get_cell_value("quantity")
                initial_quantity = 0
                if quantity is not None:
                    try:
                        initial_quantity = int(quantity)
                        product.quantity = initial_quantity
                    except (TypeError, ValueError):
                        # Ignore invalid quantity values and keep default of 0
                        pass

                low_stock_threshold = get_cell_value("low_stock_threshold")
                if low_stock_threshold is not None:
                    try:
                        product.low_stock_threshold = int(low_stock_threshold)
                    except (TypeError, ValueError):
                        # Ignore invalid threshold values and keep default
                        pass

                is_taxable = get_cell_value("is_taxable_(y/n)") or get_cell_value("is_taxable")
                if is_taxable is not None:
                    product.is_taxable = str(is_taxable).upper().strip() in ["Y", "YES", "TRUE", "1"]

                status = get_cell_value("status")
                if status is not None:
                    status_str = str(status).lower().strip()
                    if status_str == "active":
                        product.status = ProductStatus.ACTIVE
                    elif status_str == "draft":
                        product.status = ProductStatus.DRAFT
                    elif status_str == "archived":
                        product.status = ProductStatus.ARCHIVED
                else:
                    product.status = ProductStatus.ACTIVE

                if is_new:
                    self.db.add(product)
                    self.db.flush()  # Get the product ID
                    
                    # Auto-create inventory item for new products
                    if auto_create_inventory:
                        inventory_item = InventoryItem(
                            business_id=UUID(business_id),
                            product_id=product.id,
                            quantity_on_hand=initial_quantity,
                            quantity_reserved=0,
                            quantity_incoming=0,
                            reorder_point=product.low_stock_threshold or 10,
                        )
                        self.db.add(inventory_item)
                    
                    result["created"] += 1
                    if sku:
                        products_by_sku[sku] = product
                else:
                    result["updated"] += 1

                # Batch commit periodically to reduce memory pressure
                processed = result["created"] + result["updated"]
                if processed > 0 and processed % IMPORT_BATCH_SIZE == 0:
                    try:
                        self.db.commit()
                    except Exception as e:
                        self.db.rollback()
                        result["errors"].append(f"Batch commit failed at row {row_idx}: {str(e)}")

            except ValueError as e:
                result["errors"].append(f"Row {row_idx}: Invalid value - {str(e)}")
                result["skipped"] += 1
            except Exception as e:
                result["errors"].append(f"Row {row_idx}: {str(e)}")
                result["skipped"] += 1

        # Final commit
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            result["success"] = False
            result["errors"].append(f"Database error: {str(e)}")
            return result

        # Set success based on whether any items were processed
        if result["updated"] == 0 and result["created"] == 0:
            if not result["errors"]:
                result["errors"].append("No valid product data found in spreadsheet")
            result["success"] = False

        return result
