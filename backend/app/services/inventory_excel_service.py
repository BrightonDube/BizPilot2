"""Excel export/import service for inventory management."""

from io import BytesIO
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.inventory import InventoryItem
from app.models.product import Product


# Column definitions matching database schema
INVENTORY_COLUMNS = [
    {"key": "sku", "header": "SKU", "width": 15, "required": True},
    {"key": "product_name", "header": "Product Name", "width": 30, "required": False},
    {"key": "quantity_on_hand", "header": "Quantity On Hand", "width": 18, "required": True},
    {"key": "quantity_reserved", "header": "Quantity Reserved", "width": 18, "required": False},
    {"key": "quantity_incoming", "header": "Quantity Incoming", "width": 18, "required": False},
    {"key": "reorder_point", "header": "Reorder Point", "width": 15, "required": False},
    {"key": "reorder_quantity", "header": "Reorder Quantity", "width": 18, "required": False},
    {"key": "location", "header": "Location", "width": 20, "required": False},
    {"key": "bin_location", "header": "Bin Location", "width": 15, "required": False},
    {"key": "average_cost", "header": "Average Cost", "width": 15, "required": False},
    {"key": "last_cost", "header": "Last Cost", "width": 12, "required": False},
]


class InventoryExcelService:
    """Service for Excel-based inventory import/export."""

    def __init__(self, db: Session):
        self.db = db

    def _apply_header_style(self, cell: Cell) -> None:
        """Apply styling to header cells."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _apply_data_style(self, cell: Cell, is_alternate: bool = False) -> None:
        """Apply styling to data cells."""
        if is_alternate:
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_template(self) -> BytesIO:
        """Generate an empty Excel template with correct column headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Add headers
        for col_idx, col_def in enumerate(INVENTORY_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Add instructions sheet
        instructions = wb.create_sheet("Instructions")
        instructions["A1"] = "BizPilot Inventory Import Template"
        instructions["A1"].font = Font(bold=True, size=14)
        
        instructions["A3"] = "Column Descriptions:"
        instructions["A3"].font = Font(bold=True)
        
        row = 4
        for col_def in INVENTORY_COLUMNS:
            required = " (Required)" if col_def["required"] else " (Optional)"
            instructions[f"A{row}"] = f"• {col_def['header']}{required}"
            row += 1

        instructions["A" + str(row + 1)] = "Notes:"
        instructions["A" + str(row + 1)].font = Font(bold=True)
        instructions[f"A{row + 2}"] = "• SKU must match existing products in your inventory"
        instructions[f"A{row + 3}"] = "• Quantity On Hand is required for each row"
        instructions[f"A{row + 4}"] = "• Leave cells blank to keep existing values (for updates)"
        instructions[f"A{row + 5}"] = "• Costs should be numeric values without currency symbols"

        # Freeze header row
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_inventory(self, business_id: str) -> BytesIO:
        """Export all inventory items to Excel spreadsheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Add headers
        for col_idx, col_def in enumerate(INVENTORY_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Get inventory items with product info
        items = (
            self.db.query(InventoryItem, Product)
            .join(Product, InventoryItem.product_id == Product.id)
            .filter(
                InventoryItem.business_id == business_id,
                InventoryItem.deleted_at.is_(None),
                Product.deleted_at.is_(None),
            )
            .order_by(Product.sku)
            .all()
        )

        # Add data rows
        for row_idx, (item, product) in enumerate(items, 2):
            is_alternate = row_idx % 2 == 0
            
            row_data = [
                product.sku,
                product.name,
                item.quantity_on_hand,
                item.quantity_reserved,
                item.quantity_incoming,
                item.reorder_point,
                item.reorder_quantity,
                item.location,
                item.bin_location,
                float(item.average_cost) if item.average_cost else 0,
                float(item.last_cost) if item.last_cost else 0,
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_data_style(cell, is_alternate)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add metadata sheet
        meta = wb.create_sheet("Export Info")
        meta["A1"] = "Export Date:"
        meta["B1"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        meta["A2"] = "Total Items:"
        meta["B2"] = len(items)
        meta["A3"] = "Business ID:"
        meta["B3"] = str(business_id)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def import_inventory(
        self, business_id: str, file_content: bytes, user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Import inventory from Excel spreadsheet.
        
        Args:
            business_id: The business ID to import inventory for
            file_content: The Excel file content as bytes
            user_id: Optional user ID for audit tracking (reserved for future use)
        
        Returns dict with:
        - success: bool
        - updated: int (count of updated items)
        - created: int (count of new items)
        - errors: list of error messages
        - skipped: int (count of skipped rows)
        """
        result: Dict[str, Any] = {
            "success": True,
            "updated": 0,
            "created": 0,
            "errors": [],
            "skipped": 0,
        }

        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        except Exception as e:
            result["success"] = False
            result["errors"].append(f"Failed to read Excel file: {str(e)}")
            return result

        # Find the inventory sheet (handle empty workbooks)
        ws = None
        if not wb.sheetnames:
            result["success"] = False
            result["errors"].append("Excel file contains no sheets")
            return result
            
        for sheet_name in ["Inventory", "Sheet1"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        
        # Fall back to first sheet if neither found
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        if ws is None:
            result["success"] = False
            result["errors"].append("Could not find Inventory sheet")
            return result

        # Get header row and map columns
        headers: Dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header_lower = str(cell.value).lower().strip().replace(" ", "_")
                headers[header_lower] = col_idx

        # Validate required columns
        required_headers = {"sku", "quantity_on_hand"}
        missing = required_headers - set(headers.keys())
        if missing:
            result["success"] = False
            result["errors"].append(f"Missing required columns: {', '.join(missing)}")
            return result

        # Pre-fetch all products for this business to avoid N+1 queries
        products = (
            self.db.query(Product)
            .filter(
                Product.business_id == business_id,
                Product.deleted_at.is_(None),
            )
            .all()
        )
        product_by_sku: Dict[str, Product] = {p.sku: p for p in products if p.sku}
        
        # Pre-fetch all inventory items for this business
        inventory_items = (
            self.db.query(InventoryItem)
            .filter(
                InventoryItem.business_id == business_id,
                InventoryItem.deleted_at.is_(None),
            )
            .all()
        )
        inventory_by_product_id: Dict[str, InventoryItem] = {
            str(item.product_id): item for item in inventory_items
        }

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                # Get SKU
                sku_col = headers.get("sku", 1)
                sku = row[sku_col - 1].value
                
                if not sku:
                    result["skipped"] += 1
                    continue

                sku = str(sku).strip()

                # Find product by SKU (O(1) lookup from pre-fetched dict)
                product = product_by_sku.get(sku)

                if not product:
                    result["errors"].append(f"Row {row_idx}: Product with SKU '{sku}' not found")
                    result["skipped"] += 1
                    continue

                # Get or create inventory item (O(1) lookup from pre-fetched dict)
                inventory_item = inventory_by_product_id.get(str(product.id))

                is_new = inventory_item is None
                if is_new:
                    inventory_item = InventoryItem(
                        business_id=UUID(business_id),
                        product_id=product.id,
                    )
                    self.db.add(inventory_item)

                # Update fields from spreadsheet
                def get_cell_value(header_key: str, default=None):
                    col = headers.get(header_key)
                    if col is None:
                        return default
                    val = row[col - 1].value
                    return val if val is not None else default

                # Required field
                qty_on_hand = get_cell_value("quantity_on_hand")
                if qty_on_hand is None:
                    result["errors"].append(f"Row {row_idx}: Missing quantity_on_hand")
                    result["skipped"] += 1
                    continue
                
                inventory_item.quantity_on_hand = int(qty_on_hand)

                # Optional fields
                qty_reserved = get_cell_value("quantity_reserved")
                if qty_reserved is not None:
                    inventory_item.quantity_reserved = int(qty_reserved)

                qty_incoming = get_cell_value("quantity_incoming")
                if qty_incoming is not None:
                    inventory_item.quantity_incoming = int(qty_incoming)

                reorder_point = get_cell_value("reorder_point")
                if reorder_point is not None:
                    inventory_item.reorder_point = int(reorder_point)

                reorder_qty = get_cell_value("reorder_quantity")
                if reorder_qty is not None:
                    inventory_item.reorder_quantity = int(reorder_qty)

                location = get_cell_value("location")
                if location is not None:
                    # Convert empty strings to None
                    inventory_item.location = str(location).strip() or None

                bin_location = get_cell_value("bin_location")
                if bin_location is not None:
                    # Convert empty strings to None
                    inventory_item.bin_location = str(bin_location).strip() or None

                avg_cost = get_cell_value("average_cost")
                if avg_cost is not None:
                    inventory_item.average_cost = float(avg_cost)

                last_cost = get_cell_value("last_cost")
                if last_cost is not None:
                    inventory_item.last_cost = float(last_cost)

                if is_new:
                    result["created"] += 1
                    # Add to cache for subsequent lookups within batch
                    inventory_by_product_id[str(product.id)] = inventory_item
                else:
                    result["updated"] += 1

                # Batch commit every 100 rows to reduce memory pressure
                # and allow partial success on large files
                batch_size = 100
                processed = result["created"] + result["updated"]
                if processed > 0 and processed % batch_size == 0:
                    try:
                        self.db.commit()
                    except Exception as e:
                        self.db.rollback()
                        result["errors"].append(f"Batch commit failed at row {row_idx}: {str(e)}")
                        # Continue processing - already committed items are saved

            except ValueError as e:
                result["errors"].append(f"Row {row_idx}: Invalid value - {str(e)}")
                result["skipped"] += 1
            except Exception as e:
                result["errors"].append(f"Row {row_idx}: {str(e)}")
                result["skipped"] += 1

        # Final commit for remaining rows
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            result["success"] = False
            result["errors"].append(f"Database error: {str(e)}")
            return result

        # Set success based on whether any items were processed
        if result["updated"] == 0 and result["created"] == 0:
            if not result["errors"]:
                result["errors"].append("No valid inventory data found in spreadsheet")
            result["success"] = False

        return result
